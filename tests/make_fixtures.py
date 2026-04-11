"""
テスト用 Excel ファイルを生成するスクリプト。

使い方:
  python tests/make_fixtures.py
"""
import os
import sys

# パッケージを参照できるようにパスを通す
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from pathlib import Path

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def make_basic():
    """基本テスト: 行の追加・削除・セル変更が混在する例（日本語テキスト含む）"""
    # --- 旧ファイル ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [
        ["商品名",   "数量", "単価",  "合計"],
        ["りんご",   10,    100,    1000],
        ["バナナ",    5,     80,     400],   # ← 新ファイルで削除
        ["みかん",   20,    60,     1200],  # ← 単価・合計が変わる
        ["ぶどう",    3,    300,     900],
        ["いちご",    8,    200,    1600],
    ]:
        ws.append(row)
    wb.save(FIXTURES_DIR / "basic_old.xlsx")

    # --- 新ファイル ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [
        ["商品名",   "数量", "単価",  "合計"],
        ["りんご",   10,    100,    1000],
        # バナナ 削除
        ["みかん",   20,    70,     1400],  # 単価・合計 変更
        ["メロン",    2,    500,    1000],  # 新規 追加
        ["ぶどう",    3,    300,     900],
        ["いちご",    8,    200,    1600],
    ]:
        ws.append(row)
    wb.save(FIXTURES_DIR / "basic_new.xlsx")
    print("生成: basic_old.xlsx / basic_new.xlsx")


def make_multisheet():
    """複数シートテスト: シート追加・削除を含む"""
    # --- 旧ファイル ---
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "1月"
    for row in [["品目", "金額"], ["A", 100], ["B", 200], ["C", 300]]:
        ws1.append(row)
    ws2 = wb.create_sheet("2月")
    for row in [["品目", "金額"], ["A", 150], ["B", 250]]:
        ws2.append(row)
    ws3 = wb.create_sheet("削除シート")
    ws3.append(["このシートは削除されます"])
    wb.save(FIXTURES_DIR / "multi_old.xlsx")

    # --- 新ファイル ---
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "1月"
    for row in [["品目", "金額"], ["A", 100], ["B", 220], ["C", 300]]:  # B 変更
        ws1.append(row)
    ws2 = wb.create_sheet("2月")
    for row in [["品目", "金額"], ["A", 150], ["B", 250], ["D", 400]]:  # D 追加
        ws2.append(row)
    ws4 = wb.create_sheet("追加シート")  # 削除シート → 追加シート
    ws4.append(["このシートは追加されました"])
    wb.save(FIXTURES_DIR / "multi_new.xlsx")
    print("生成: multi_old.xlsx / multi_new.xlsx")


def make_no_diff():
    """差分なしテスト: 完全に同一のファイル"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [["名前", "値"], ["A", 1], ["B", 2], ["C", 3]]:
        ws.append(row)
    wb.save(FIXTURES_DIR / "nodiff_old.xlsx")
    wb.save(FIXTURES_DIR / "nodiff_new.xlsx")
    print("生成: nodiff_old.xlsx / nodiff_new.xlsx")


def make_matcher_fixture():
    """カスタムマッチャーテスト用: コード改番が含まれる例"""
    # --- 旧ファイル ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "取引"
    for row in [
        ["取引先名",    "旧コード", "金額"],
        ["A商事",      "CODE001", 10000],
        ["B物産",      "CODE002", 20000],
        ["C株式会社",  "CODE003", 30000],  # ← コードは変わるが金額も変わる（差分あり）
        ["D商会",      "CODE999", 50000],  # ← マッピング外（差分として検出）
    ]:
        ws.append(row)
    wb.save(FIXTURES_DIR / "matcher_old.xlsx")

    # --- 新ファイル（コード改番後）---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "取引"
    for row in [
        ["取引先名",    "新コード", "金額"],
        ["A商事",      "NEW001",  10000],  # CODE001→NEW001: マッチャーでEQUAL
        ["B物産",      "NEW002",  20000],  # CODE002→NEW002: マッチャーでEQUAL
        ["C株式会社",  "NEW003",  99999],  # CODE003→NEW003はEQUAL、金額変更は差分
        ["D商会",      "UNKNOWN", 50000],  # マッピング外の変換 → 差分あり
    ]:
        ws.append(row)
    wb.save(FIXTURES_DIR / "matcher_new.xlsx")

    # --- マッチャー設定 JSON ---
    import json
    config = [
        {
            "type": "mapping",
            "column": "B",
            "sheet": None,
            "pairs": [
                ["CODE001", "NEW001"],
                ["CODE002", "NEW002"],
                ["CODE003", "NEW003"],
            ]
        }
    ]
    with open(FIXTURES_DIR / "matchers.json", "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print("生成: matcher_old.xlsx / matcher_new.xlsx / matchers.json")


def make_chartest():
    """
    文字レベルdiff の視覚確認用フィクスチャ。
    商品マスタ形式（5列）で、行追加・削除・複数列変更を含む。

    変更の種類:
      数値1桁変更     : みかん 単価 60→70
      全角末尾追加    : みかん 備考 国産→国産（新物）
      行削除          : バナナ 行ごと削除
      全角中間変更    : 東京牛乳→大阪牛乳（東京→大阪）
      全角部分変更×2  : 東京都産→大阪府産 / 山田（東京）→鈴木（大阪）
      数値+半角変更   : 1本500ml→1本1000ml / 単価120→130
      変更なし        : チョコレート 行まるごと同一
      行挿入          : クリームパン 新規
      半角1文字変更   : 6枚切り→8枚切り
      半角バージョン  : v1.0.0→v1.2.0 / 初回限定→通常版
    """
    header = ["コード", "商品名", "カテゴリ", "単価", "備考", "担当者"]

    old_data = [
        # コード      商品名              カテゴリ       単価  備考           担当者
        ["A001",  "みかん",          "果物",        60,   "国産",        "山田（東京）"],
        ["A002",  "バナナ",          "果物",        80,   "輸入品",      "田中"],           # ← 削除
        ["A003",  "東京牛乳",        "乳製品",      150,  "東京都産",    "山田（東京）"],   # ← 複数列変更
        ["A004",  "りんごジュース",  "飲料",        120,  "1本500ml",    "鈴木"],           # ← 単価+備考変更
        ["A005",  "チョコレート",    "菓子",        200,  "ベルギー産",  "鈴木"],           # ← 変更なし
        ["A006",  "食パン",          "パン",        180,  "6枚切り",     "田中"],           # ← 備考変更
        ["A007",  "v1.0.0ソフト",   "ソフトウェア", 9800, "初回限定",   "管理者"],         # ← 商品名+備考変更
    ]

    new_data = [
        # コード      商品名              カテゴリ       単価  備考           担当者
        ["A001",  "みかん",          "果物",        70,   "国産（新物）", "山田（東京）"],  # 単価+備考変更
        # A002 削除
        ["A003",  "大阪牛乳",        "乳製品",      150,  "大阪府産",    "鈴木（大阪）"],  # 商品名+備考+担当変更
        ["A004",  "りんごジュース",  "飲料",        130,  "1本1000ml",   "鈴木"],          # 単価+備考変更
        ["A005",  "チョコレート",    "菓子",        200,  "ベルギー産",  "鈴木"],          # 変更なし
        ["A008",  "クリームパン",    "パン",        220,  "あんこ入り",  "田中"],          # 新規追加
        ["A006",  "食パン",          "パン",        180,  "8枚切り",     "田中"],          # 備考変更
        ["A007",  "v1.2.0ソフト",   "ソフトウェア", 9800, "通常版",     "管理者"],        # 商品名+備考変更
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品マスタ"
    ws.append(header)
    for row in old_data:
        ws.append(row)
    wb.save(FIXTURES_DIR / "chartest_old.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品マスタ"
    ws.append(header)
    for row in new_data:
        ws.append(row)
    wb.save(FIXTURES_DIR / "chartest_new.xlsx")

    print(f"生成: chartest_old.xlsx / chartest_new.xlsx"
          f" (旧{len(old_data)}行 → 新{len(new_data)}行, 5列)")


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    make_basic()
    make_multisheet()
    make_no_diff()
    make_matcher_fixture()
    make_chartest()
    print(f"\nすべてのテストデータを生成しました: {FIXTURES_DIR}/")

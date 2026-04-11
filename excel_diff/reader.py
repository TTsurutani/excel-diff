"""
Excelファイルの読み込みモジュール。
openpyxl を使ってシート・行・セルのデータを取り出す。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class CellData:
    value: Any
    strikethrough: bool = False

    def display(self) -> str:
        if self.value is None:
            return ""
        return str(self.value)


@dataclass
class RowData:
    row_idx: int          # 元ファイルでの行番号（1始まり）
    cells: list[CellData] = field(default_factory=list)

    def hash_key(self, include_strikethrough: bool = False) -> tuple:
        """LCS用のハッシュキーを返す（カスタムマッチャーは含まない）。"""
        if include_strikethrough:
            return tuple((c.value, c.strikethrough) for c in self.cells)
        return tuple(c.value for c in self.cells)


@dataclass
class SheetData:
    name: str
    rows: list[RowData]
    max_col: int


def read_workbook(
    path: str,
    include_strikethrough: bool = False,
    sheet_filter: Optional[str] = None,
) -> dict[str, SheetData]:
    """
    Excelファイルを読み込み {シート名: SheetData} を返す。

    Parameters
    ----------
    path:
        読み込む .xlsx ファイルのパス
    include_strikethrough:
        True のとき取り消し線情報を CellData に格納する
    sheet_filter:
        指定時はそのシートのみ読み込む
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, SheetData] = {}

    for ws in wb.worksheets:
        if sheet_filter and ws.title != sheet_filter:
            continue

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        rows: list[RowData] = []
        for r_idx in range(1, max_row + 1):
            cells: list[CellData] = []
            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                strike = False
                if include_strikethrough:
                    try:
                        strike = bool(cell.font and cell.font.strike)
                    except Exception:
                        strike = False
                cells.append(CellData(value=cell.value, strikethrough=strike))
            rows.append(RowData(row_idx=r_idx, cells=cells))

        # 末尾の空行（全セル None）を除去
        while rows and all(c.value is None for c in rows[-1].cells):
            rows.pop()

        result[ws.title] = SheetData(name=ws.title, rows=rows, max_col=max_col)

    wb.close()
    return result

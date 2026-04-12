"""
差分アルゴリズムモジュール。

行レベルのLCS（SequenceMatcher）→ セルレベルdiff の2段階で差分を計算する。
カスタムマッチャーはLCS用の行ハッシュ正規化とセル比較の両方に適用される。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from enum import Enum
from typing import Any, Optional

from openpyxl.utils import get_column_letter

from .reader import CellData, RowData, SheetData
from .matcher import ColumnMatcher


# ---------------------------------------------------------------------------
# データモデル
# ---------------------------------------------------------------------------

class RowTag(Enum):
    EQUAL  = "equal"
    DELETE = "delete"
    INSERT = "insert"
    MODIFY = "modify"


@dataclass
class CellDiff:
    col_idx: int               # 0始まり列インデックス
    old_cell: Optional[CellData]
    new_cell: Optional[CellData]


@dataclass
class RowDiff:
    tag: RowTag
    old_row: Optional[RowData] = None
    new_row: Optional[RowData] = None
    cell_diffs: list[CellDiff] = field(default_factory=list)  # MODIFY のときのみ


@dataclass
class SheetDiff:
    name: str
    status: str                           # "equal" / "modified" / "added" / "deleted"
    row_diffs: list[RowDiff] = field(default_factory=list)
    max_cols: int = 0
    col_letters: list[str] = field(default_factory=list)   # ["A", "B", ...]
    col_filter: Optional[set[int]] = None  # 比較対象列（None = 全列）


@dataclass
class FileDiff:
    old_path: str
    new_path: str
    sheet_diffs: list[SheetDiff] = field(default_factory=list)
    has_differences: bool = False
    matcher_count: int = 0           # 適用されたマッチャー数（HTML表示用）


# ---------------------------------------------------------------------------
# 内部ユーティリティ
# ---------------------------------------------------------------------------

def _pad_cells(cells: list[CellData], target_len: int) -> list[CellData]:
    """セルリストを target_len 長にパディングして返す（元リストは変更しない）。"""
    if len(cells) >= target_len:
        return cells[:target_len]
    return cells + [CellData(None)] * (target_len - len(cells))


def _normalize_val(v: Any) -> Any:
    """
    セル値を比較用に正規化する。
    - None と空文字列は同一視（None に統一）
    - '_x000D_' 形式のテキスト埋め込み CR 表現を除去
    - CR+LF / CR のみ → LF に統一（改行コード差異を吸収）
    """
    if v is None:
        return None
    if isinstance(v, str):
        v = re.sub(r'_?[xX]000[dD]_?', '', v)  # _x000D_ または x000D を除去
        v = v.replace('\r\n', '\n')               # CRLF → LF
        v = v.replace('\r', '\n')                 # CR のみ → LF
    return None if v == "" else v


def _cell_equal(
    old_cell: CellData,
    new_cell: CellData,
    col_idx: int,
    sheet_name: str,
    matchers: list[ColumnMatcher],
    include_strike: bool,
) -> bool:
    """2つのセルが「等値」かどうかを判定する。"""
    # カスタムマッチャーが適用される列かチェック
    for m in matchers:
        if m.applies_to(sheet_name, col_idx):
            return m.matches(old_cell.value, new_cell.value)

    # デフォルト比較（None と空文字列は同一視）
    if include_strike:
        return (
            _normalize_val(old_cell.value) == _normalize_val(new_cell.value)
            and old_cell.strikethrough == new_cell.strikethrough
        )
    return _normalize_val(old_cell.value) == _normalize_val(new_cell.value)


def _normalize_row_key(
    cells: list[CellData],
    side: str,   # "old" or "new"
    sheet_name: str,
    matchers: list[ColumnMatcher],
    include_strike: bool,
    col_filter: Optional[set[int]],
) -> tuple:
    """
    行をLCS比較用の正規化キーに変換する。
    col_filter が指定されている場合は対象列のみハッシュに含める。
    カスタムマッチャーが適用される列は (_SENTINEL, canonical_old_val) に正規化される。
    """
    key: list[Any] = []
    for i, cell in enumerate(cells):
        if col_filter is not None and i not in col_filter:
            continue  # 比較対象外の列はハッシュから除外
        val = cell.value
        matched = False
        for m in matchers:
            if m.applies_to(sheet_name, i):
                if side == "old":
                    val = m.normalize_old(cell.value)
                else:
                    val = m.normalize_new(cell.value)
                matched = True
                break
        if not matched and include_strike:
            val = (val, cell.strikethrough)
        key.append(val)
    return tuple(key)


def _row_similarity(
    row_a: RowData,
    row_b: RowData,
    max_cols: int,
    col_filter: Optional[set[int]],
) -> float:
    """
    2行の類似度を返す（比較対象列における一致セル数 / 比較対象列数）。
    col_filter が None の場合は全列を対象とする。
    """
    a_cells = _pad_cells(row_a.cells, max_cols)
    b_cells = _pad_cells(row_b.cells, max_cols)
    cols = range(max_cols) if col_filter is None else sorted(col_filter)
    total = len(cols)
    if total == 0:
        return 0.0
    matches = sum(
        1 for i in cols
        if i < len(a_cells) and i < len(b_cells) and a_cells[i].value == b_cells[i].value
    )
    return matches / total


def _pair_replace_rows(
    a_rows: list[RowData],
    b_rows: list[RowData],
    max_cols: int,
    col_filter: Optional[set[int]],
) -> list[tuple]:
    """
    replace ブロック内の行を類似度でペアリングする。
    比較対象列で 1セルも一致しないペアは作らず、残りは DELETE / INSERT として返す。
    戻り値: [(old_row|None, new_row|None), ...]
    """
    if not a_rows:
        return [(None, b) for b in b_rows]
    if not b_rows:
        return [(a, None) for a in a_rows]

    scored = sorted(
        [(_row_similarity(a_rows[i], b_rows[j], max_cols, col_filter), i, j)
         for i in range(len(a_rows))
         for j in range(len(b_rows))],
        reverse=True,
    )

    pairs: list[tuple] = []
    used_a: set[int] = set()
    used_b: set[int] = set()

    for score, i, j in scored:
        if i in used_a or j in used_b:
            continue
        if score <= 0.0:
            break
        pairs.append((a_rows[i], b_rows[j]))
        used_a.add(i)
        used_b.add(j)

    for i, a in enumerate(a_rows):
        if i not in used_a:
            pairs.append((a, None))
    for j, b in enumerate(b_rows):
        if j not in used_b:
            pairs.append((None, b))

    return pairs


def _compute_cell_diffs(
    old_row: RowData,
    new_row: RowData,
    max_cols: int,
    sheet_name: str,
    matchers: list[ColumnMatcher],
    include_strike: bool,
    col_filter: Optional[set[int]],
) -> list[CellDiff]:
    """MODIFY行のセルレベル差分を計算する（比較対象列のみ）。"""
    old_cells = _pad_cells(old_row.cells, max_cols)
    new_cells = _pad_cells(new_row.cells, max_cols)
    diffs: list[CellDiff] = []
    for i, (oc, nc) in enumerate(zip(old_cells, new_cells)):
        if col_filter is not None and i not in col_filter:
            continue  # 比較対象外の列はスキップ
        if not _cell_equal(oc, nc, i, sheet_name, matchers, include_strike):
            diffs.append(CellDiff(col_idx=i, old_cell=oc, new_cell=nc))
    return diffs


# ---------------------------------------------------------------------------
# メイン差分ロジック
# ---------------------------------------------------------------------------

def _diff_sheet_rows(
    sheet_name: str,
    rows_a: list[RowData],
    rows_b: list[RowData],
    max_cols: int,
    matchers: list[ColumnMatcher],
    include_strike: bool,
    col_filter: Optional[set[int]],
) -> list[RowDiff]:
    """行リスト同士をLCSで差分計算し RowDiff のリストを返す。"""

    keys_a = [
        _normalize_row_key(_pad_cells(r.cells, max_cols), "old", sheet_name, matchers, include_strike, col_filter)
        for r in rows_a
    ]
    keys_b = [
        _normalize_row_key(_pad_cells(r.cells, max_cols), "new", sheet_name, matchers, include_strike, col_filter)
        for r in rows_b
    ]

    matcher = SequenceMatcher(None, keys_a, keys_b, autojunk=False)
    result: list[RowDiff] = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for ra, rb in zip(rows_a[i1:i2], rows_b[j1:j2]):
                result.append(RowDiff(RowTag.EQUAL, ra, rb))

        elif tag == "delete":
            for ra in rows_a[i1:i2]:
                result.append(RowDiff(RowTag.DELETE, ra, None))

        elif tag == "insert":
            for rb in rows_b[j1:j2]:
                result.append(RowDiff(RowTag.INSERT, None, rb))

        elif tag == "replace":
            a_slice = rows_a[i1:i2]
            b_slice = rows_b[j1:j2]

            for old_row, new_row in _pair_replace_rows(a_slice, b_slice, max_cols, col_filter):
                if old_row is None:
                    result.append(RowDiff(RowTag.INSERT, None, new_row))
                elif new_row is None:
                    result.append(RowDiff(RowTag.DELETE, old_row, None))
                else:
                    cell_diffs = _compute_cell_diffs(
                        old_row, new_row, max_cols, sheet_name, matchers, include_strike, col_filter
                    )
                    if cell_diffs:
                        result.append(RowDiff(RowTag.MODIFY, old_row, new_row, cell_diffs))
                    else:
                        result.append(RowDiff(RowTag.EQUAL, old_row, new_row))

    return result


def diff_files(
    old_sheets: dict[str, SheetData],
    new_sheets: dict[str, SheetData],
    old_path: str,
    new_path: str,
    matchers: Optional[list[ColumnMatcher]] = None,
    include_strike: bool = False,
    config: Optional[object] = None,   # DiffConfig（循環import回避のため型はobject）
) -> FileDiff:
    """
    2つのブック（シート辞書）の差分を計算して FileDiff を返す。

    Parameters
    ----------
    old_sheets / new_sheets:
        reader.read_workbook() の戻り値
    old_path / new_path:
        表示用ファイルパス
    matchers:
        カスタムマッチャーのリスト（config未指定時に使用）
    include_strike:
        取り消し線を差分として扱うか
    config:
        DiffConfig オブジェクト（matchers + 列フィルタをまとめて指定）
    """
    # config が渡された場合はそちらを優先
    if config is not None:
        effective_matchers = config.matchers
        get_col_filter = config.get_col_filter
    else:
        effective_matchers = matchers or []
        get_col_filter = lambda _sheet: None  # noqa: E731

    result = FileDiff(
        old_path=old_path,
        new_path=new_path,
        matcher_count=len(effective_matchers),
    )

    all_names: list[str] = list(dict.fromkeys(list(old_sheets.keys()) + list(new_sheets.keys())))

    for name in all_names:
        old_s = old_sheets.get(name)
        new_s = new_sheets.get(name)
        col_filter = get_col_filter(name)

        if old_s is None:
            max_cols = new_s.max_col
            col_letters = [get_column_letter(i) for i in range(1, max_cols + 1)]
            row_diffs = [RowDiff(RowTag.INSERT, None, r) for r in new_s.rows]
            result.sheet_diffs.append(SheetDiff(name, "added", row_diffs, max_cols, col_letters, col_filter))
            result.has_differences = True

        elif new_s is None:
            max_cols = old_s.max_col
            col_letters = [get_column_letter(i) for i in range(1, max_cols + 1)]
            row_diffs = [RowDiff(RowTag.DELETE, r, None) for r in old_s.rows]
            result.sheet_diffs.append(SheetDiff(name, "deleted", row_diffs, max_cols, col_letters, col_filter))
            result.has_differences = True

        else:
            max_cols = max(old_s.max_col, new_s.max_col, 1)
            col_letters = [get_column_letter(i) for i in range(1, max_cols + 1)]
            sheet_matchers = [m for m in effective_matchers if m.sheet is None or m.sheet == name]

            row_diffs = _diff_sheet_rows(
                name,
                list(old_s.rows),
                list(new_s.rows),
                max_cols,
                sheet_matchers,
                include_strike,
                col_filter,
            )
            has_diff = any(rd.tag != RowTag.EQUAL for rd in row_diffs)
            status = "modified" if has_diff else "equal"
            result.sheet_diffs.append(SheetDiff(name, status, row_diffs, max_cols, col_letters, col_filter))
            if has_diff:
                result.has_differences = True

    return result

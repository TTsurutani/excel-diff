"""
excel-diff CLI エントリポイント。

使い方:
  # ファイル比較
  python -m excel_diff old.xlsx new.xlsx
  python -m excel_diff old.xlsx new.xlsx -o diff.html --open

  # フォルダ一括比較（完全一致）
  python -m excel_diff --dir old_dir/ new_dir/

  # フォルダ一括比較（ペアJSON使用）
  python -m excel_diff --dir old_dir/ new_dir/ --pairs pairs.json

  # フォルダ一括比較（パターン使用）
  python -m excel_diff --dir old_dir/ new_dir/ --pattern monthly

  # ペア候補探索
  python -m excel_diff --discover old_dir/ new_dir/ -o pairs.json

  # ペアからパターン生成
  python -m excel_diff --gen-pattern pairs.json --id monthly --name "月次レポート"

  # パターン一覧
  python -m excel_diff --list-patterns

  # ブックをシート単位に分解
  python -m excel_diff --split book.xlsx --prefix "2024_"
  python -m excel_diff --split book.xlsx --suffix "_output" --output-dir out/
"""
from __future__ import annotations

import argparse
import os
import sys
import webbrowser
from pathlib import Path

from excel_diff.utils import generate_output_dir


# ---------------------------------------------------------------------------
# 引数パーサ
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="excel-diff",
        description="Excelファイルの差分をHTMLで出力します",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
例:
  excel-diff old.xlsx new.xlsx
  excel-diff old.xlsx new.xlsx -o diff.html --open
  excel-diff --dir old_dir new_dir
  excel-diff --dir old_dir new_dir --pairs pairs.json
  excel-diff --dir old_dir new_dir --pattern monthly
  excel-diff --discover old_dir new_dir -o pairs.json
  excel-diff --gen-pattern pairs.json --id monthly --name "月次レポート"
  excel-diff --list-patterns
""",
    )

    # --- ファイル比較 ---
    p.add_argument("old_file", nargs="?", help="比較元ファイル (.xlsx)")
    p.add_argument("new_file", nargs="?", help="比較先ファイル (.xlsx)")
    p.add_argument("-o", "--output", metavar="PATH",
                   help="出力HTMLパス（省略時: <新ファイル名>_diff.html）")

    # --- フォルダ比較 ---
    p.add_argument("--dir", nargs=2, metavar=("OLD_DIR", "NEW_DIR"),
                   help="フォルダ一括比較")
    p.add_argument("--output-dir", metavar="DIR",
                   help="フォルダ比較時の出力先ディレクトリ")

    # --- ペア探索 ---
    p.add_argument("--discover", nargs=2, metavar=("OLD_DIR", "NEW_DIR"),
                   help="ファイルペア候補を探索してJSONに保存")
    p.add_argument("--threshold", type=float, default=0.3, metavar="SCORE",
                   help="--discover の類似度しきい値 (0.0〜1.0, デフォルト: 0.3)")

    # --- パターン生成 ---
    p.add_argument("--gen-pattern", metavar="PAIRS_JSON",
                   help="確認済みペアJSONからパターン（正規表現）を生成・保存")
    p.add_argument("--id", metavar="ID",
                   help="--gen-pattern: パターンID")
    p.add_argument("--name", metavar="NAME",
                   help="--gen-pattern: パターン名")
    p.add_argument("--regex", metavar="REGEX",
                   help="--gen-pattern: 正規表現を手動指定（自動生成をスキップ）")
    p.add_argument("--desc", metavar="TEXT",
                   help="--gen-pattern: パターンの説明")
    p.add_argument("--example-old", metavar="DIR",
                   help="--gen-pattern: 旧フォルダの例")
    p.add_argument("--example-new", metavar="DIR",
                   help="--gen-pattern: 新フォルダの例")

    # --- パターン一覧 ---
    p.add_argument("--list-patterns", action="store_true",
                   help="保存済みパターンを一覧表示")

    # --- ブック分解 ---
    p.add_argument("--split", metavar="FILE",
                   help="Excelブックをシート単位のファイルに分解する")
    p.add_argument("--prefix", metavar="TEXT", default="",
                   help="--split: 出力ファイル名の前置文字列")
    p.add_argument("--suffix", metavar="TEXT", default="",
                   help="--split: 出力ファイル名の後置文字列（拡張子の前）")
    p.add_argument("--name-regex", metavar="PATTERN", default=None,
                   help="--split: シート名からファイル名ベースを抽出する正規表現（第1キャプチャグループを使用）。例: \"^([^（]+)\" で和名のみ抽出")

    # --- パターン指定（フォルダ比較時） ---
    p.add_argument("--pairs", metavar="FILE",
                   help="フォルダ比較時に使用するペアJSONファイル（--discover で生成したもの）")
    p.add_argument("--pattern", metavar="ID",
                   help="フォルダ比較時に使用するパターンID")
    p.add_argument("--patterns-file", metavar="FILE", default="patterns.json",
                   help="パターン定義ファイルのパス (デフォルト: patterns.json)")

    # --- 共通オプション ---
    p.add_argument("--sheet", metavar="NAME",
                   help="指定シートのみ比較・完全一致（省略時: 全シート）")
    p.add_argument("--sheet-old", metavar="PATTERN",
                   help="old側シートを正規表現で指定（部分一致）。複数マッチ時は先頭を使用")
    p.add_argument("--sheet-new", metavar="PATTERN",
                   help="new側シートを正規表現で指定（部分一致）。複数マッチ時は先頭を使用")
    p.add_argument("--strikethrough", action="store_true",
                   help="取り消し線の有無も差分として扱う")
    p.add_argument("--matchers", metavar="FILE",
                   help="カスタムマッチャー設定JSONファイル")
    p.add_argument("--include-cols", metavar="SPEC",
                   help="比較対象列（例: A:C,E）。省略時は全列比較")
    p.add_argument("--open", dest="open", action="store_true", default=True,
                   help="生成後にブラウザで自動オープン（デフォルト: ON）")
    p.add_argument("--no-open", dest="open", action="store_false",
                   help="生成後にブラウザを自動オープンしない")
    p.add_argument("--key-cols", metavar="SPEC",
                   help="キーJOIN差分モードのキー列（例: B,C）。指定するだけで key モードが有効になる")
    p.add_argument("--diff-mode", choices=["lcs", "key"], default=None,
                   help="差分モード: lcs（出現順LCS、デフォルト）または key（キーJOIN）")
    return p


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def _default_output_path(new_file: str) -> str:
    stem = Path(new_file).stem
    parent = Path(new_file).parent
    return str(parent / f"{stem}_diff.html")


def _diff_stats(file_diff) -> tuple[int, int, int]:
    """(削除行数, 追加行数, 変更行数) を返す。"""
    from .diff_engine import RowTag
    delete = sum(1 for sd in file_diff.sheet_diffs for rd in sd.row_diffs if rd.tag == RowTag.DELETE)
    insert = sum(1 for sd in file_diff.sheet_diffs for rd in sd.row_diffs if rd.tag == RowTag.INSERT)
    modify = sum(1 for sd in file_diff.sheet_diffs for rd in sd.row_diffs if rd.tag == RowTag.MODIFY)
    return delete, insert, modify


def _stats_line(delete: int, insert: int, modify: int) -> str:
    parts = []
    if delete: parts.append(f"削除 {delete}行")
    if insert: parts.append(f"追加 {insert}行")
    if modify: parts.append(f"変更 {modify}行")
    return "、".join(parts) if parts else "行変更なし"


def _write_index_xlsx(
    results: list,
    unmatched: list,
    old_dir: str,
    new_dir: str,
    out_path: str,
) -> None:
    """フォルダ比較のインデックスを ★index.xlsx として書き出す。"""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
    from openpyxl.utils import get_column_letter

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    no_diff = [(pair, fd, op) for pair, fd, op in results if not fd.has_differences]
    has_diff = [(pair, fd, op) for pair, fd, op in results if fd.has_differences]
    old_label = Path(old_dir).name or old_dir
    new_label = Path(new_dir).name or new_dir

    wb = Workbook()
    ws = wb.active
    ws.title = "比較結果"

    # ---- スタイル定数 ----
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT   = Font(bold=True, size=13)
    META_FONT    = Font(size=10, color="57606A")
    CARD_DIFF    = PatternFill("solid", fgColor="FFEEF0")
    CARD_NODIFF  = PatternFill("solid", fgColor="E6FFED")
    CARD_TOTAL   = PatternFill("solid", fgColor="F6F8FA")
    NUM_DIFF_DEL = Font(color="CF222E", bold=True)
    NUM_DIFF_INS = Font(color="1A7F37", bold=True)
    NUM_DIFF_MOD = Font(color="9A6700", bold=True)
    NUM_DIFF_TOT = Font(bold=True)
    NODIFF_FONT  = Font(color="8B949E")
    LINK_FONT    = Font(color="0000CD", underline="single")
    UNMATCH_FONT = Font(color="57606A", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="D0D7DE")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_border(cell):
        cell.border = BORDER

    # ---- タイトル行 ----
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "excel-diff 比較結果インデックス"
    c.font  = TITLE_FONT
    c.alignment = LEFT
    ws.row_dimensions[1].height = 22

    ws["A2"].value = f"旧: {old_dir}"
    ws["A2"].font  = META_FONT
    ws.merge_cells("A2:H2")
    ws["A3"].value = f"新: {new_dir}"
    ws["A3"].font  = META_FONT
    ws.merge_cells("A3:H3")
    ws["A4"].value = f"生成: {now}"
    ws["A4"].font  = META_FONT
    ws.merge_cells("A4:H4")

    # ---- サマリカード（行6-8, 列A-H を3列に分割） ----
    ws.row_dimensions[5].height = 8  # spacer
    card_data = [
        ("差分あり", len(has_diff), CARD_DIFF,  Font(color="CF222E", bold=True, size=18)),
        ("差分なし", len(no_diff),  CARD_NODIFF, Font(color="1A7F37", bold=True, size=18)),
        ("合計",     len(results),  CARD_TOTAL,  Font(bold=True, size=18)),
    ]
    col_map = [(1, 3), (4, 6), (7, 8)]  # (開始列, 終了列) ※1始まり
    for (label, count, fill, num_font), (cs, ce) in zip(card_data, col_map):
        cs_letter = get_column_letter(cs)
        ce_letter = get_column_letter(ce)
        top_ref    = f"{cs_letter}6:{ce_letter}6"
        bottom_ref = f"{cs_letter}7:{ce_letter}7"
        ws.merge_cells(top_ref)
        ws.merge_cells(bottom_ref)
        num_cell = ws[f"{cs_letter}6"]
        lbl_cell = ws[f"{cs_letter}7"]
        num_cell.value = count
        num_cell.font  = num_font
        num_cell.fill  = fill
        num_cell.alignment = CENTER
        lbl_cell.value = label
        lbl_cell.font  = Font(size=10)
        lbl_cell.fill  = fill
        lbl_cell.alignment = CENTER
        for r in (6, 7):
            for col in range(cs, ce + 1):
                ws.cell(row=r, column=col).fill = fill
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 8  # spacer

    # ---- 注釈 ----
    ws.merge_cells("A9:H9")
    note = ws["A9"]
    note.value = "※ 行数はファイル内の全シートを合算した値です。シート別の内訳は各差分HTMLを参照してください。"
    note.font  = Font(size=9, color="57606A")

    # ---- テーブルヘッダ ----
    headers = ["キー", old_label, new_label, "削除", "追加", "変更", "合計", "リンク"]
    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=10, column=col_idx, value=hdr)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = RIGHT if col_idx >= 4 else LEFT
        set_border(c)
    ws.row_dimensions[10].height = 18

    # ---- データ行 ----
    data_row = 11
    for pair, file_diff, out_file in has_diff:
        delete, insert, modify = _diff_stats(file_diff)
        total_changes = delete + insert + modify
        rel = Path(out_file).name
        cells_vals = [
            (pair.key or "",  LEFT,  None),
            (pair.old_name,   LEFT,  None),
            (pair.new_name,   LEFT,  None),
            (delete,          RIGHT, NUM_DIFF_DEL),
            (insert,          RIGHT, NUM_DIFF_INS),
            (modify,          RIGHT, NUM_DIFF_MOD),
            (total_changes,   RIGHT, NUM_DIFF_TOT),
            ("開く",          CENTER, LINK_FONT),
        ]
        for col_idx, (val, align, fnt) in enumerate(cells_vals, 1):
            c = ws.cell(row=data_row, column=col_idx, value=val)
            c.alignment = align
            if fnt:
                c.font = fnt
            set_border(c)
        # ハイパーリンク
        link_cell = ws.cell(row=data_row, column=8)
        link_cell.hyperlink = rel
        link_cell.font = LINK_FONT
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    for pair, file_diff, out_file in no_diff:
        rel = Path(out_file).name
        cells_vals = [
            (pair.key or "",  LEFT,   NODIFF_FONT),
            (pair.old_name,   LEFT,   NODIFF_FONT),
            (pair.new_name,   LEFT,   NODIFF_FONT),
            ("−",             CENTER, NODIFF_FONT),
            ("−",             CENTER, NODIFF_FONT),
            ("−",             CENTER, NODIFF_FONT),
            ("−",             CENTER, NODIFF_FONT),
            ("開く",          CENTER, LINK_FONT),
        ]
        for col_idx, (val, align, fnt) in enumerate(cells_vals, 1):
            c = ws.cell(row=data_row, column=col_idx, value=val)
            c.alignment = align
            if fnt:
                c.font = fnt
            set_border(c)
        link_cell = ws.cell(row=data_row, column=8)
        link_cell.hyperlink = rel
        link_cell.font = LINK_FONT
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # ---- 比較対象外 ----
    if unmatched:
        data_row += 1
        ws.merge_cells(f"A{data_row}:H{data_row}")
        h = ws[f"A{data_row}"]
        h.value = "比較対象外"
        h.font  = Font(bold=True, size=11)
        data_row += 1
        for p in unmatched:
            if p.old_name and not p.new_name:
                label = f"[旧のみ] {p.old_name}"
            elif p.new_name and not p.old_name:
                label = f"[新のみ] {p.new_name}"
            else:
                continue
            ws.merge_cells(f"A{data_row}:H{data_row}")
            c = ws[f"A{data_row}"]
            c.value = label
            c.font  = UNMATCH_FONT
            data_row += 1

    # ---- 列幅 ----
    col_widths = [15, 40, 40, 8, 8, 8, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ウィンドウ固定（ヘッダ行） ----
    ws.freeze_panes = "A11"

    wb.save(out_path)


def _render_index_html(
    results: list,
    unmatched: list,
    old_dir: str,
    new_dir: str,
) -> str:
    """フォルダ比較のインデックスHTMLを生成して返す。

    行数統計（削除・追加・変更）はファイル内の全シートを合算した値。
    シート別の内訳は各差分HTMLを参照すること。
    """
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    no_diff = [(pair, fd, op) for pair, fd, op in results if not fd.has_differences]
    has_diff = [(pair, fd, op) for pair, fd, op in results if fd.has_differences]

    old_label = Path(old_dir).name or old_dir
    new_label = Path(new_dir).name or new_dir

    rows_html = ""
    for pair, file_diff, out_path in has_diff:
        delete, insert, modify = _diff_stats(file_diff)
        total_changes = delete + insert + modify
        rel_path = Path(out_path).name
        rows_html += (
            f'<tr class="has-diff">'
            f'<td class="filename">{pair.old_name}</td>'
            f'<td class="filename">{pair.new_name}</td>'
            f'<td class="num del">{delete}</td>'
            f'<td class="num ins">{insert}</td>'
            f'<td class="num mod">{modify}</td>'
            f'<td class="num total">{total_changes}</td>'
            f'<td class="link"><a href="{rel_path}" target="_blank">開く</a></td>'
            f'</tr>\n'
        )

    for pair, file_diff, out_path in no_diff:
        rel_path = Path(out_path).name
        rows_html += (
            f'<tr class="no-diff">'
            f'<td class="filename">{pair.old_name}</td>'
            f'<td class="filename">{pair.new_name}</td>'
            f'<td class="num">−</td>'
            f'<td class="num">−</td>'
            f'<td class="num">−</td>'
            f'<td class="num">−</td>'
            f'<td class="link"><a href="{rel_path}" target="_blank">開く</a></td>'
            f'</tr>\n'
        )

    unmatched_html = ""
    if unmatched:
        unmatched_html = "<h2>比較対象外</h2><ul>"
        for p in unmatched:
            if p.old_name and not p.new_name:
                unmatched_html += f"<li>[旧のみ] {p.old_name}</li>"
            elif p.new_name and not p.old_name:
                unmatched_html += f"<li>[新のみ] {p.new_name}</li>"
        unmatched_html += "</ul>"

    total = len(results)
    diff_count = len(has_diff)
    nodiff_count = len(no_diff)

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>excel-diff インデックス</title>
<style>
  body {{ font-family: system-ui, sans-serif; margin: 0; background: #f6f8fa; color: #24292f; }}
  .topbar {{ background: linear-gradient(mediumblue, darkblue); color: #fff; padding: 10px 20px; }}
  .topbar h1 {{ margin: 0; font-size: 18px; }}
  .topbar .meta {{ font-size: 12px; opacity: .8; margin-top: 4px; }}
  .container {{ max-width: 1100px; margin: 24px auto; padding: 0 16px; }}
  .summary {{ display: flex; gap: 16px; margin-bottom: 20px; }}
  .summary-card {{ background: #fff; border: 1px solid #d0d7de; border-radius: 6px;
                   padding: 12px 20px; flex: 1; text-align: center; }}
  .summary-card .num {{ font-size: 28px; font-weight: bold; }}
  .summary-card.diff .num {{ color: #cf222e; }}
  .summary-card.nodiff .num {{ color: #1a7f37; }}
  .note {{ font-size: 12px; color: #57606a; margin-bottom: 12px; }}
  .table-wrapper {{ border: 1px solid #d0d7de; border-radius: 6px; overflow: hidden;
                    max-height: calc(100vh - 260px); overflow-y: auto; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff; }}
  thead th {{ position: sticky; top: 0; z-index: 1;
              background: #f6f8fa; padding: 8px 12px;
              border-bottom: 2px solid #d0d7de;
              text-align: left; font-size: 12px; color: #57606a; white-space: nowrap; }}
  td {{ padding: 7px 12px; border-bottom: 1px solid #f0f0f0; font-size: 13px; }}
  tr:last-child td {{ border-bottom: none; }}
  td.num {{ text-align: right; width: 52px; white-space: nowrap; }}
  tr.has-diff td.del {{ color: #cf222e; font-weight: bold; }}
  tr.has-diff td.ins {{ color: #1a7f37; font-weight: bold; }}
  tr.has-diff td.mod {{ color: #9a6700; font-weight: bold; }}
  tr.has-diff td.total {{ color: #24292f; font-weight: bold; border-left: 1px solid #d0d7de; }}
  tr.no-diff {{ color: #8b949e; }}
  td.link {{ width: 48px; text-align: center; }}
  td.link a {{ color: mediumblue; text-decoration: none; font-weight: bold; }}
  td.link a:hover {{ text-decoration: underline; }}
  h2 {{ font-size: 15px; margin-top: 24px; }}
  ul {{ font-size: 13px; color: #57606a; }}
</style>
</head>
<body>
<div class="topbar">
  <h1>excel-diff 比較結果インデックス</h1>
  <div class="meta">
    旧: {old_dir} &nbsp;→&nbsp; 新: {new_dir}<br>
    生成: {now}
  </div>
</div>
<div class="container">
  <div class="summary">
    <div class="summary-card diff"><div class="num">{diff_count}</div><div>差分あり</div></div>
    <div class="summary-card nodiff"><div class="num">{nodiff_count}</div><div>差分なし</div></div>
    <div class="summary-card"><div class="num">{total}</div><div>合計</div></div>
  </div>
  <p class="note">※ 行数はファイル内の全シートを合算した値です。シート別の内訳は各差分HTMLを参照してください。</p>
  <div class="table-wrapper">
    <table>
      <thead>
        <tr>
          <th>{old_label}</th>
          <th>{new_label}</th>
          <th style="text-align:right">削除</th>
          <th style="text-align:right">追加</th>
          <th style="text-align:right">変更</th>
          <th style="text-align:right">合計</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
{rows_html}      </tbody>
    </table>
  </div>
  {unmatched_html}
</div>
</body>
</html>"""


def _build_config(args: argparse.Namespace):
    """引数から DiffConfig を組み立てて返す。"""
    from .matcher import load_config, DiffConfig, parse_col_spec, parse_col_list

    if args.matchers:
        if not os.path.isfile(args.matchers):
            print(f"エラー: マッチャー設定ファイルが見つかりません: {args.matchers}", file=sys.stderr)
            sys.exit(1)
        config = load_config(args.matchers)
        print(f"カスタムマッチャー: {len(config.matchers)} 件ロード")
    else:
        config = DiffConfig()

    if args.include_cols:
        config.global_col_filter = parse_col_spec(args.include_cols)

    # --key-cols が指定された場合は自動的に key モードへ
    if args.key_cols:
        config.key_cols = parse_col_list(args.key_cols)
        config.diff_mode = "key"

    # --diff-mode が明示指定された場合はそちらを優先
    if args.diff_mode:
        config.diff_mode = args.diff_mode

    if config.diff_mode == "key":
        if not config.key_cols:
            print("エラー: --diff-mode key を使うには --key-cols でキー列を指定してください", file=sys.stderr)
            sys.exit(1)
        from openpyxl.utils import get_column_letter
        cols_disp = ", ".join(get_column_letter(c + 1) for c in config.key_cols)
        print(f"差分モード: key JOIN  キー列: {cols_disp}")

    return config


def _collect_pairs_exact(old_dir: str, new_dir: str):
    """完全一致ベースのペアリング（--pattern 未指定時）。"""
    from .file_pairing import FilePair
    old_files = {
        f for f in os.listdir(old_dir)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    }
    new_files = {
        f for f in os.listdir(new_dir)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    }
    all_names = sorted(old_files | new_files)
    return [
        FilePair(
            old_name=name if name in old_files else None,
            new_name=name if name in new_files else None,
            score=1.0,
            matched_by="exact",
        )
        for name in all_names
    ]


# ---------------------------------------------------------------------------
# --discover
# ---------------------------------------------------------------------------

def _run_discover(args: argparse.Namespace) -> None:
    from .file_pairing import discover_pairs, save_pairs

    old_dir, new_dir = args.discover

    for d in (old_dir, new_dir):
        if not os.path.isdir(d):
            print(f"エラー: ディレクトリが見つかりません: {d}", file=sys.stderr)
            sys.exit(1)

    pairs = discover_pairs(old_dir, new_dir, threshold=args.threshold)

    print("ペア候補:")
    for p in pairs:
        if p.matched_by == "exact":
            print(f"  [完全一致] {p.old_name}")
        elif p.matched_by == "auto":
            print(f"  [自動 {p.score:.0%}] {p.old_name}  →  {p.new_name}")
        elif p.matched_by == "unmatched_old":
            print(f"  [未対応-旧] {p.old_name}")
        elif p.matched_by == "unmatched_new":
            print(f"  [未対応-新] {p.new_name}")

    output = args.output or "pairs.json"
    save_pairs(pairs, output)
    print(f"\n→ {output} に保存しました")
    print("内容を確認・編集後、--gen-pattern で正規表現を生成してください。")


# ---------------------------------------------------------------------------
# --gen-pattern
# ---------------------------------------------------------------------------

def _run_gen_pattern(args: argparse.Namespace) -> None:
    from datetime import date
    from .file_pairing import load_pairs, generate_regex, validate_regex
    from .patterns import PatternStore, PatternDef

    if not args.id or not args.name:
        print("エラー: --id と --name は必須です", file=sys.stderr)
        sys.exit(1)

    pairs_path = args.gen_pattern
    if not os.path.isfile(pairs_path):
        print(f"エラー: ペアファイルが見つかりません: {pairs_path}", file=sys.stderr)
        sys.exit(1)

    pairs = load_pairs(pairs_path)
    matched = [p for p in pairs if p.old_name and p.new_name]

    if not matched:
        print("エラー: 比較可能なペアがありません", file=sys.stderr)
        sys.exit(1)

    # 正規表現の決定
    if args.regex:
        key_regex = args.regex
        print(f"指定正規表現: {key_regex}")
    else:
        key_regex = generate_regex(pairs)
        if key_regex:
            print(f"提案パターン: {key_regex}")
            print()
            print("キー抽出例:")
            import re
            pat = re.compile(key_regex)
            for p in matched[:5]:  # 最大5件表示
                m_old = pat.fullmatch(p.old_name)
                m_new = pat.fullmatch(p.new_name)
                key = m_old.group(1) if m_old else "?"
                print(f"  {p.old_name}  →  キー: \"{key}\"")
                print(f"  {p.new_name}  →  キー: \"{m_new.group(1) if m_new else '?'}\"")
                print()
        else:
            print("正規表現の自動生成に失敗しました。")
            print("--regex オプションで正規表現を手動指定してください。")
            print(r'例: --regex "^(.+?)_\d{8}\.xlsx$"')
            sys.exit(1)

    # 検証
    print("検証中...")
    errors = validate_regex(matched, key_regex)

    if errors:
        print("[NG] 検証失敗:")
        for e in errors:
            print(f"  [{e.kind}] {e.details}")
        print()
        print("正規表現を修正して --regex オプションで再指定してください。")
        sys.exit(1)

    print(f"[OK] 検証OK: {len(matched)} ペアすべてが正しく再現されます")

    # パターン保存
    store = PatternStore(args.patterns_file)
    pattern = PatternDef(
        id=args.id,
        name=args.name,
        key_regex=key_regex,
        description=args.desc or "",
        example_old_dir=args.example_old or "",
        example_new_dir=args.example_new or "",
        created_at=str(date.today()),
    )
    store.add_or_update(pattern)
    store.save()

    print()
    print(f"パターンを保存しました: {args.patterns_file}")
    print(f"  ID      : {pattern.id}")
    print(f"  名前    : {pattern.name}")
    print(f"  正規表現: {pattern.key_regex}")
    print()
    print(f"使用方法: excel-diff --dir old/ new/ --pattern {pattern.id}")


# ---------------------------------------------------------------------------
# --list-patterns
# ---------------------------------------------------------------------------

def _run_list_patterns(args: argparse.Namespace) -> None:
    from .patterns import PatternStore

    store = PatternStore(args.patterns_file)
    patterns = store.list_all()

    if not patterns:
        print(f"保存済みパターンはありません。({args.patterns_file})")
        return

    print(f"保存済みパターン一覧: {args.patterns_file}")
    for p in patterns:
        print()
        print(f"  ID      : {p.id}")
        print(f"  名前    : {p.name}")
        print(f"  正規表現: {p.key_regex}")
        if p.description:
            print(f"  説明    : {p.description}")
        if p.example_old_dir:
            print(f"  例 (旧) : {p.example_old_dir}")
        if p.example_new_dir:
            print(f"  例 (新) : {p.example_new_dir}")
        if p.created_at:
            print(f"  作成日  : {p.created_at}")


# ---------------------------------------------------------------------------
# ブック分解
# ---------------------------------------------------------------------------

def _run_split(args: argparse.Namespace) -> None:
    from .splitter import split_workbook

    path = args.split
    if not os.path.isfile(path):
        print(f"エラー: ファイルが見つかりません: {path}", file=sys.stderr)
        sys.exit(1)
    if not path.lower().endswith(".xlsx"):
        print(f"エラー: .xlsx ファイルを指定してください: {path}", file=sys.stderr)
        sys.exit(1)

    prefix = args.prefix
    suffix = args.suffix
    name_regex = args.name_regex
    output_dir = args.output_dir  # None の場合はブックと同フォルダ

    print(f"分解中: {path}")
    if prefix:
        print(f"  前置: {prefix}")
    if suffix:
        print(f"  後置: {suffix}")
    if name_regex:
        print(f"  名前変換正規表現: {name_regex}")

    try:
        output_paths = split_workbook(
            path, prefix=prefix, suffix=suffix, name_regex=name_regex, output_dir=output_dir
        )
    except ValueError as e:
        print(f"エラー: {e}", file=sys.stderr)
        sys.exit(1)

    print()
    print(f"{len(output_paths)} シートを分解しました:")
    for p in output_paths:
        print(f"  → {p}")


# ---------------------------------------------------------------------------
# ファイル比較
# ---------------------------------------------------------------------------

def _run_file_diff(args: argparse.Namespace) -> None:
    from .reader import read_workbook, filter_sheets_by_pattern
    from .diff_engine import diff_files
    from .html_renderer import render

    old_path = args.old_file
    new_path = args.new_file

    if not os.path.isfile(old_path):
        print(f"エラー: ファイルが見つかりません: {old_path}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(new_path):
        print(f"エラー: ファイルが見つかりません: {new_path}", file=sys.stderr)
        sys.exit(1)

    config = _build_config(args)

    sheet_old_pat = getattr(args, "sheet_old", None)
    sheet_new_pat = getattr(args, "sheet_new", None)

    print(f"読み込み中: {old_path}")
    old_sheets = read_workbook(old_path, args.strikethrough, args.sheet)
    if sheet_old_pat:
        old_sheets = filter_sheets_by_pattern(old_sheets, sheet_old_pat)
        if not old_sheets:
            print(f"エラー: --sheet-old '{sheet_old_pat}' にマッチするシートがありません: {old_path}", file=sys.stderr)
            sys.exit(1)

    print(f"読み込み中: {new_path}")
    new_sheets = read_workbook(new_path, args.strikethrough, args.sheet)
    if sheet_new_pat:
        new_sheets = filter_sheets_by_pattern(new_sheets, sheet_new_pat)
        if not new_sheets:
            print(f"エラー: --sheet-new '{sheet_new_pat}' にマッチするシートがありません: {new_path}", file=sys.stderr)
            sys.exit(1)

    print("差分計算中...")
    file_diff = diff_files(
        old_sheets, new_sheets,
        old_path, new_path,
        include_strike=args.strikethrough,
        config=config,
    )

    output_path = args.output or _default_output_path(new_path)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(render(file_diff))

    print()
    if file_diff.has_differences:
        delete, insert, modify = _diff_stats(file_diff)
        print(f"差分なし: 0 ファイル")
        print(f"差分あり: 1 ファイル")
        print(f"  {Path(new_path).name}  ({_stats_line(delete, insert, modify)})  → {output_path}")
    else:
        print(f"差分なし: 1 ファイル")
        print(f"差分あり: 0 ファイル")

    if args.open:
        webbrowser.open(Path(output_path).resolve().as_uri())


# ---------------------------------------------------------------------------
# フォルダ比較
# ---------------------------------------------------------------------------

def _run_dir_diff(args: argparse.Namespace) -> None:
    from .reader import read_workbook, filter_sheets_by_pattern
    from .diff_engine import diff_files
    from .html_renderer import render
    from .file_pairing import apply_pattern

    old_dir, new_dir = args.dir

    for d in (old_dir, new_dir):
        if not os.path.isdir(d):
            print(f"エラー: ディレクトリが見つかりません: {d}", file=sys.stderr)
            sys.exit(1)

    config = _build_config(args)

    # ペアリング（--pairs > --pattern > 完全一致 の優先順）
    if args.pairs:
        from .file_pairing import load_pairs
        if not os.path.isfile(args.pairs):
            print(f"エラー: ペアファイルが見つかりません: {args.pairs}", file=sys.stderr)
            sys.exit(1)
        pairs = load_pairs(args.pairs)
        print(f"ペアファイル使用: {args.pairs}  ({len([p for p in pairs if p.old_name and p.new_name])} ペア)")
    elif args.pattern:
        from .patterns import PatternStore
        store = PatternStore(args.patterns_file)
        pat = store.get(args.pattern)
        if pat is None:
            print(f"エラー: パターン '{args.pattern}' が見つかりません ({args.patterns_file})", file=sys.stderr)
            sys.exit(1)
        print(f"パターン「{pat.name}」を使用")
        print(f"  正規表現: {pat.key_regex}")
        pairs = apply_pattern(old_dir, new_dir, pat.key_regex)

        # パターンペアリング結果を表示
        named_pairs = [p for p in pairs if p.old_name and p.new_name and p.old_name != p.new_name]
        if named_pairs:
            print("ペアリング:")
            for p in named_pairs:
                print(f"  {p.old_name}  →  {p.new_name}")
    else:
        pairs = _collect_pairs_exact(old_dir, new_dir)

    compare_pairs = [p for p in pairs if p.old_name and p.new_name]
    unmatched = [p for p in pairs if not p.old_name or not p.new_name]

    if not compare_pairs:
        print("比較対象の .xlsx ファイルが見つかりませんでした。")
        return

    # 出力ディレクトリ
    if args.output_dir:
        out_dir = args.output_dir
    else:
        out_dir = generate_output_dir(old_dir, new_dir)

    os.makedirs(out_dir, exist_ok=True)
    print(f"出力先: {out_dir}/")

    sheet_old_pat = getattr(args, "sheet_old", None)
    sheet_new_pat = getattr(args, "sheet_new", None)

    # 比較処理
    results = []
    for pair in compare_pairs:
        old_path = os.path.join(old_dir, pair.old_name)
        new_path = os.path.join(new_dir, pair.new_name)

        try:
            old_sheets = read_workbook(old_path, args.strikethrough, args.sheet) if os.path.isfile(old_path) else {}
            if sheet_old_pat and old_sheets:
                old_sheets = filter_sheets_by_pattern(old_sheets, sheet_old_pat)
                if not old_sheets:
                    print(f"  警告: --sheet-old '{sheet_old_pat}' にマッチするシートなし: {pair.old_name}  スキップします")
                    continue
        except Exception as e:
            print(f"  警告: 旧ファイルを読み込めません（空として扱います）: {pair.old_name}  ({e})")
            old_sheets = {}
        try:
            new_sheets = read_workbook(new_path, args.strikethrough, args.sheet) if os.path.isfile(new_path) else {}
            if sheet_new_pat and new_sheets:
                new_sheets = filter_sheets_by_pattern(new_sheets, sheet_new_pat)
                if not new_sheets:
                    print(f"  警告: --sheet-new '{sheet_new_pat}' にマッチするシートなし: {pair.new_name}  スキップします")
                    continue
        except Exception as e:
            print(f"  警告: 新ファイルを読み込めません（空として扱います）: {pair.new_name}  ({e})")
            new_sheets = {}

        file_diff = diff_files(
            old_sheets, new_sheets,
            old_path, new_path,
            include_strike=args.strikethrough,
            config=config,
        )

        # 出力HTMLのファイル名は新ファイルのステムを使う
        stem = Path(pair.new_name).stem
        out_path = os.path.join(out_dir, f"{stem}_diff.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(render(file_diff))

        results.append((pair, file_diff, out_path))

    # サマリ出力
    no_diff = [r for r in results if not r[1].has_differences]
    has_diff = [r for r in results if r[1].has_differences]

    print()
    print(f"差分なし: {len(no_diff)} ファイル")
    print(f"差分あり: {len(has_diff)} ファイル")
    for pair, file_diff, out_path in has_diff:
        delete, insert, modify = _diff_stats(file_diff)
        label = pair.new_name
        if pair.old_name != pair.new_name:
            label = f"{pair.old_name} → {pair.new_name}"
        print(f"  {label}  ({_stats_line(delete, insert, modify)})  → {out_path}")

    if unmatched:
        print()
        print(f"比較対象外: {len(unmatched)} ファイル")
        for p in unmatched:
            if p.old_name and not p.new_name:
                print(f"  [旧のみ] {p.old_name}")
            elif p.new_name and not p.old_name:
                print(f"  [新のみ] {p.new_name}")

    # インデックスHTMLを生成
    # index_path = os.path.join(out_dir, "★index.html")
    # with open(index_path, "w", encoding="utf-8") as f:
    #     f.write(_render_index_html(results, unmatched, old_dir, new_dir))
    # print(f"\nインデックス → {index_path}")

    # インデックスXLSXを生成
    index_xlsx_path = os.path.join(out_dir, "★index.xlsx")
    _write_index_xlsx(results, unmatched, old_dir, new_dir, index_xlsx_path)
    print(f"インデックス → {index_xlsx_path}")

    if args.open:
        os.startfile(index_xlsx_path)


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    if args.split:
        _run_split(args)
    elif args.discover:
        _run_discover(args)
    elif args.gen_pattern:
        _run_gen_pattern(args)
    elif args.list_patterns:
        _run_list_patterns(args)
    elif args.dir:
        _run_dir_diff(args)
    elif args.old_file and args.new_file:
        _run_file_diff(args)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()

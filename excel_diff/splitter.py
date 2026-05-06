"""
Excelブックをシート単位のファイルに分解するモジュール。

split_workbook(path, prefix, suffix, name_regex, output_dir) → list[str]
  各シートを <output_dir>/<prefix><ファイル名ベース><suffix>.xlsx として保存し、
  出力ファイルパスのリストを返す。

Windows 環境では win32com (Excel COM) を優先使用し、データ入力規則・条件付き書式・
ピボットテーブルなどを保持したまま分割する。win32com が利用できない場合は
openpyxl にフォールバックする。
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path


# ファイル名として使えない文字（Windows / macOS / Linux 共通の危険文字）
_INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')


def _safe_filename(sheet_name: str) -> str:
    """シート名をファイル名として安全な文字列に変換する。"""
    return _INVALID_CHARS.sub("_", sheet_name)


def _apply_name_regex(sheet_name: str, pattern: re.Pattern[str]) -> str:
    """
    正規表現の第1キャプチャグループにマッチした部分を返す。
    マッチしない場合はシート名全体にフォールバックして警告を出す。
    """
    m = pattern.search(sheet_name)
    if m and m.lastindex and m.lastindex >= 1:
        return m.group(1)
    warnings.warn(
        f"--name-regex がシート '{sheet_name}' にマッチしませんでした。シート名をそのまま使用します。",
        stacklevel=3,
    )
    return sheet_name


def _build_output_paths(
    sheet_names: list[str],
    prefix: str,
    suffix: str,
    compiled_regex: re.Pattern[str] | None,
    out_dir: Path,
) -> list[Path]:
    paths = []
    for sheet_name in sheet_names:
        if compiled_regex is not None:
            name_base = _apply_name_regex(sheet_name, compiled_regex)
        else:
            name_base = sheet_name
        safe_name = _safe_filename(name_base)
        paths.append(out_dir / f"{prefix}{safe_name}{suffix}.xlsx")
    return paths


def _split_via_com(
    src_path: Path,
    sheet_names: list[str],
    out_paths: list[Path],
) -> None:
    """win32com (Excel COM) を使ってシート分割する。"""
    import pythoncom
    import win32com.client

    # ワーカースレッドから呼ばれる場合に備えて COM を初期化
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        for sheet_name, out_path in zip(sheet_names, out_paths):
            wb = excel.Workbooks.Open(str(src_path))
            try:
                # 削除前に対象シートを表示状態にする（xlSheetVisible = -1）
                # ※先に可視化しないと「可視シートがゼロ」エラーになる
                ws = wb.Sheets(sheet_name)
                if ws.Visible != -1:
                    ws.Visible = -1

                # 対象シート以外を削除
                for ws in list(wb.Sheets):
                    if ws.Name != sheet_name:
                        ws.Delete()

                # 他シート参照の名前定義を削除して #REF! エラーを防止
                for dn in list(wb.Names):
                    dn.Delete()

                # xlOpenXMLWorkbook = 51 で .xlsx として保存
                wb.SaveAs(str(out_path), FileFormat=51)
            finally:
                wb.Close(False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


def _split_via_openpyxl(
    src_path: Path,
    sheet_names: list[str],
    out_paths: list[Path],
) -> None:
    """openpyxl を使ってシート分割する（win32com 非対応環境用フォールバック）。"""
    from openpyxl import load_workbook

    for sheet_name, out_path in zip(sheet_names, out_paths):
        wb = load_workbook(str(src_path))
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]

        ws = wb[sheet_name]
        if ws.sheet_state == 'hidden':
            ws.sheet_state = 'visible'

        for dn in list(wb.defined_names):
            del wb.defined_names[dn]

        wb.save(str(out_path))
        wb.close()


def split_workbook(
    path: str,
    prefix: str = "",
    suffix: str = "",
    name_regex: str | None = None,
    output_dir: str | None = None,
) -> list[str]:
    """
    ブックを1シート1ファイルに分解して保存する。

    Parameters
    ----------
    path       : 入力Excelファイルパス (.xlsx)
    prefix     : 出力ファイル名の前置文字列
    suffix     : 出力ファイル名の後置文字列（拡張子の前）
    name_regex : ファイル名ベース抽出用正規表現（第1キャプチャグループを使用）。
                 Noneの場合はシート名をそのまま使用。
    output_dir : 出力先ディレクトリ（Noneの場合はブックと同じフォルダ）

    Returns
    -------
    出力ファイルパスのリスト（シート順）
    """
    compiled_regex: re.Pattern[str] | None = None
    if name_regex:
        compiled_regex = re.compile(name_regex)
        if compiled_regex.groups < 1:
            raise ValueError(
                f"--name-regex にはキャプチャグループ () が1つ以上必要です: {name_regex!r}"
            )

    src_path = Path(path).resolve()
    out_dir = Path(output_dir).resolve() if output_dir else src_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # シート名を取得
    try:
        from openpyxl import load_workbook as _lw
        wb_meta = _lw(str(src_path), read_only=True, data_only=True)
        sheet_names = wb_meta.sheetnames
        wb_meta.close()
    except ImportError as e:
        raise ImportError("openpyxl が必要です: pip install openpyxl") from e

    out_paths = _build_output_paths(sheet_names, prefix, suffix, compiled_regex, out_dir)

    # win32com を優先、利用不可なら openpyxl にフォールバック
    try:
        import win32com.client  # noqa: F401
        import pythoncom        # noqa: F401
        _split_via_com(src_path, sheet_names, out_paths)
    except ImportError:
        _split_via_openpyxl(src_path, sheet_names, out_paths)

    return [str(p) for p in out_paths]

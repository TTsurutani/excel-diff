"""タブ② フォルダ比較。"""
import os
import queue
import tkinter as tk
import webbrowser
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Callable

from . import settings as cfg
from .widgets import FileSelectRow
from .worker import get_worker


class TabDirDiff(tk.Frame):

    def __init__(self, parent, log: Callable[[str], None]) -> None:
        super().__init__(parent)
        self._log = log
        self._result_q: "queue.Queue | None" = None

        self._old     = tk.StringVar(value=cfg.get("dir_diff", "old_dir"))
        self._new     = tk.StringVar(value=cfg.get("dir_diff", "new_dir"))
        self._pairing = tk.StringVar(value=cfg.get("dir_diff", "pairing", "exact"))
        self._pairs_f = tk.StringVar(value=cfg.get("dir_diff", "pairs_file"))
        self._pat_id  = tk.StringVar(value=cfg.get("dir_diff", "pattern_id"))
        self._out_dir = tk.StringVar(value=cfg.get("dir_diff", "output_dir"))
        self._sheet   = tk.StringVar(value=cfg.get("dir_diff", "sheet"))
        self._cols    = tk.StringVar(value=cfg.get("dir_diff", "include_cols"))
        self._matchers= tk.StringVar(value=cfg.get("dir_diff", "matchers"))
        self._key_cols= tk.StringVar(value=cfg.get("dir_diff", "key_cols"))
        self._strike  = tk.BooleanVar(value=cfg.get("dir_diff", "strikethrough"))
        self._open_br = tk.BooleanVar(value=cfg.get("dir_diff", "open_browser", True))
        self._mode    = tk.StringVar(value=cfg.get("dir_diff", "diff_mode", "lcs"))

        self._opt_built = False
        self._opt_open  = False
        self._patterns: list = []

        self._build()

    # ------------------------------------------------------------------ レイアウト

    def _build(self) -> None:
        pad = {"padx": 6, "pady": 3}

        # フォルダ選択
        grp_dirs = tk.LabelFrame(self, text="フォルダ")
        grp_dirs.pack(fill="x", **pad)
        FileSelectRow(grp_dirs, "旧フォルダ", self._old, mode="dir").pack(
            fill="x", padx=6, pady=2)
        FileSelectRow(grp_dirs, "新フォルダ", self._new, mode="dir").pack(
            fill="x", padx=6, pady=2)

        # ペアリング方法
        grp_pair = tk.LabelFrame(self, text="ペアリング方法")
        grp_pair.pack(fill="x", **pad)

        tk.Radiobutton(
            grp_pair, text="完全一致（同名ファイルを対応付ける・デフォルト）",
            variable=self._pairing, value="exact", command=self._on_pairing,
        ).pack(anchor="w", padx=8)

        fr_pairs = tk.Frame(grp_pair)
        fr_pairs.pack(anchor="w", padx=8, pady=(0, 2))
        tk.Radiobutton(
            fr_pairs, text="ペアJSON",
            variable=self._pairing, value="pairs", command=self._on_pairing,
        ).pack(side="left")
        tk.Label(fr_pairs, text="ファイル").pack(side="left", padx=(8, 2))
        self._entry_pairs = tk.Entry(fr_pairs, textvariable=self._pairs_f, width=28)
        self._entry_pairs.pack(side="left")
        self._btn_pairs = tk.Button(
            fr_pairs, text="参照", width=6,
            command=self._browse_pairs,
        )
        self._btn_pairs.pack(side="left", padx=2)

        fr_pat = tk.Frame(grp_pair)
        fr_pat.pack(anchor="w", padx=8, pady=(0, 4))
        tk.Radiobutton(
            fr_pat, text="パターン",
            variable=self._pairing, value="pattern", command=self._on_pairing,
        ).pack(side="left")
        tk.Label(fr_pat, text="パターン名").pack(side="left", padx=(8, 2))
        self._cmb_pat = ttk.Combobox(fr_pat, textvariable=self._pat_id, width=24, state="disabled")
        self._cmb_pat.pack(side="left")
        tk.Button(
            fr_pat, text="更新", width=5, command=self._reload_patterns,
        ).pack(side="left", padx=2)

        # 差分モード
        grp_mode = tk.LabelFrame(self, text="差分モード")
        grp_mode.pack(fill="x", **pad)

        tk.Radiobutton(
            grp_mode, text="LCS（行の出現順で比較・デフォルト）",
            variable=self._mode, value="lcs", command=self._on_mode,
        ).pack(anchor="w", padx=8)

        fr_key = tk.Frame(grp_mode)
        fr_key.pack(anchor="w", padx=8, pady=(0, 4))
        tk.Radiobutton(
            fr_key, text="キーJOIN（キー列の値で行を対応付ける）",
            variable=self._mode, value="key", command=self._on_mode,
        ).pack(side="left")
        tk.Label(fr_key, text="キー列").pack(side="left", padx=(12, 2))
        self._entry_key = tk.Entry(fr_key, textvariable=self._key_cols, width=18)
        self._entry_key.pack(side="left")
        tk.Label(fr_key, text="例: C  または  B,C", fg="gray").pack(side="left", padx=6)

        # オプション折りたたみ + 実行ボタン（同じ行）
        ctrl_row = tk.Frame(self)
        ctrl_row.pack(fill="x", padx=6, pady=(6, 2))

        self._lbl_opt = tk.Label(ctrl_row, text="▶ オプション", cursor="hand2", fg="blue")
        self._lbl_opt.pack(side="left")
        self._lbl_opt.bind("<Button-1>", self._toggle_opt)

        self._btn_run = tk.Button(
            ctrl_row, text="実行", width=16,
            bg="#4a9eff", fg="white", font=("", 10, "bold"),
            command=self._run,
        )
        self._btn_run.pack(side="right")

        # オプション本体（初期は非表示）
        self._grp_opt = tk.LabelFrame(self, text="オプション")

        self._on_pairing()
        self._on_mode()
        self._reload_patterns()

    def _browse_pairs(self) -> None:
        from tkinter import filedialog
        path = filedialog.askopenfilename(filetypes=[("JSON", "*.json"), ("All", "*.*")])
        if path:
            self._pairs_f.set(path)

    def _reload_patterns(self) -> None:
        try:
            from excel_diff.patterns import PatternStore
            store = PatternStore(cfg.patterns_file())
            self._patterns = store.list_all()
        except Exception:
            self._patterns = []
        values = [f"{p.id}  {p.name}" for p in self._patterns]
        self._cmb_pat["values"] = values
        if self._patterns and not self._pat_id.get():
            self._pat_id.set(values[0])

    def _on_pairing(self) -> None:
        method = self._pairing.get()
        pairs_state = "normal" if method == "pairs" else "disabled"
        pat_state   = "readonly" if method == "pattern" else "disabled"
        self._entry_pairs.config(state=pairs_state)
        self._btn_pairs.config(state=pairs_state)
        self._cmb_pat.config(state=pat_state)

    def _on_mode(self) -> None:
        state = "normal" if self._mode.get() == "key" else "disabled"
        self._entry_key.config(state=state)

    def _toggle_opt(self, _=None) -> None:
        if self._opt_open:
            self._grp_opt.pack_forget()
            self._opt_open = False
            self._lbl_opt.config(text="▶ オプション")
        else:
            if not self._opt_built:
                self._build_opt()
                self._opt_built = True
            self._grp_opt.pack(fill="x", padx=6, pady=3)
            self._opt_open = True
            self._lbl_opt.config(text="▼ オプション")

    def _build_opt(self) -> None:
        pad = {"padx": 6, "pady": 2}
        g = self._grp_opt

        FileSelectRow(g, "出力フォルダ", self._out_dir, mode="dir").pack(fill="x", **pad)

        fr = tk.Frame(g)
        fr.pack(fill="x", **pad)
        tk.Label(fr, text="比較シート", width=14, anchor="w").pack(side="left")
        tk.Entry(fr, textvariable=self._sheet).pack(side="left", fill="x", expand=True)
        tk.Label(fr, text="空=全シート", fg="gray").pack(side="left", padx=4)

        fr2 = tk.Frame(g)
        fr2.pack(fill="x", **pad)
        tk.Label(fr2, text="比較列", width=14, anchor="w").pack(side="left")
        tk.Entry(fr2, textvariable=self._cols).pack(side="left", fill="x", expand=True)
        tk.Label(fr2, text="例: A:C,E", fg="gray").pack(side="left", padx=4)

        FileSelectRow(
            g, "マッチャーJSON", self._matchers,
            filetypes=[("JSON", "*.json"), ("All", "*.*")],
        ).pack(fill="x", **pad)

        tk.Checkbutton(
            g, text="取り消し線も差分として扱う", variable=self._strike,
        ).pack(anchor="w", **pad)
        tk.Checkbutton(
            g, text="完了後ブラウザで開く", variable=self._open_br,
        ).pack(anchor="w", **pad)

    # ------------------------------------------------------------------ 実行

    def _run(self) -> None:
        old = self._old.get().strip()
        new = self._new.get().strip()
        if not old or not new:
            messagebox.showerror("エラー", "旧フォルダと新フォルダを指定してください")
            return
        if not os.path.isdir(old):
            messagebox.showerror("エラー", f"フォルダが見つかりません:\n{old}")
            return
        if not os.path.isdir(new):
            messagebox.showerror("エラー", f"フォルダが見つかりません:\n{new}")
            return

        method = self._pairing.get()
        if method == "pairs" and not self._pairs_f.get().strip():
            messagebox.showerror("エラー", "ペアJSONファイルを指定してください")
            return
        if method == "pattern":
            pat_sel = self._pat_id.get().strip()
            pat_id = pat_sel.split()[0] if pat_sel else ""
            if not pat_id:
                messagebox.showerror("エラー", "パターンを選択してください")
                return
        else:
            pat_id = ""

        if self._mode.get() == "key" and not self._key_cols.get().strip():
            messagebox.showerror("エラー", "キーJOINモード: キー列を指定してください")
            return

        cfg.set_tab("dir_diff", {
            "old_dir": old, "new_dir": new,
            "pairing": method,
            "pairs_file": self._pairs_f.get(),
            "pattern_id": pat_id,
            "output_dir": self._out_dir.get(),
            "sheet": self._sheet.get(),
            "include_cols": self._cols.get(),
            "matchers": self._matchers.get(),
            "strikethrough": self._strike.get(),
            "open_browser": self._open_br.get(),
            "diff_mode": self._mode.get(),
            "key_cols": self._key_cols.get(),
        })
        cfg.save()

        self._btn_run.config(state="disabled", text="実行中...")
        self._log("フォルダ比較を開始します...")

        self._result_q = get_worker().submit(
            self._do_diff,
            old, new, method,
            self._pairs_f.get().strip(), pat_id,
            self._out_dir.get().strip(), self._sheet.get().strip(),
            self._cols.get().strip(), self._matchers.get().strip(),
            self._strike.get(), self._mode.get(),
            self._key_cols.get().strip(), self._open_br.get(),
        )
        self.after(100, self._poll)

    def _poll(self) -> None:
        if self._result_q is None:
            return
        try:
            status, val = self._result_q.get_nowait()
            self._btn_run.config(state="normal", text="実行")
            if status == "err":
                self._log(f"エラー: {val}")
        except queue.Empty:
            self.after(100, self._poll)

    def _do_diff(
        self,
        old_dir, new_dir, pairing_method,
        pairs_file, pattern_id,
        output_dir, sheet, include_cols,
        matchers_file, strikethrough, diff_mode, key_cols_str, open_browser,
    ) -> None:
        from excel_diff.reader import read_workbook
        from excel_diff.diff_engine import diff_files
        from excel_diff.html_renderer import render
        from excel_diff.matcher import DiffConfig, parse_col_spec, parse_col_list, load_config
        from excel_diff.file_pairing import FilePair
        from openpyxl.utils import get_column_letter
        from excel_diff.__main__ import _render_index_html

        # ペアリング
        if pairing_method == "pairs":
            self._log(f"ペアJSON読み込み: {pairs_file}")
            from excel_diff.file_pairing import load_pairs
            pairs = load_pairs(pairs_file)
        elif pairing_method == "pattern":
            self._log(f"パターン適用: {pattern_id}")
            from excel_diff.patterns import PatternStore
            from excel_diff.file_pairing import apply_pattern
            store = PatternStore(cfg.patterns_file())
            pat = store.get(pattern_id)
            if pat is None:
                raise ValueError(f"パターンが見つかりません: {pattern_id}")
            pairs = apply_pattern(old_dir, new_dir, pat.key_regex)
        else:
            self._log("完全一致ペアリング...")
            old_files = {
                f for f in os.listdir(old_dir)
                if f.lower().endswith(".xlsx") and not f.startswith("~$")
            }
            new_files = {
                f for f in os.listdir(new_dir)
                if f.lower().endswith(".xlsx") and not f.startswith("~$")
            }
            all_names = sorted(old_files | new_files)
            pairs = [
                FilePair(
                    old_name=name if name in old_files else None,
                    new_name=name if name in new_files else None,
                    score=1.0, matched_by="exact",
                )
                for name in all_names
            ]

        matched   = [p for p in pairs if p.old_name and p.new_name]
        unmatched = [p for p in pairs if not p.old_name or not p.new_name]
        self._log(f"ペア: 比較対象 {len(matched)} 件、対象外 {len(unmatched)} 件")

        if not matched:
            self._log("比較対象のペアがありません")
            return

        # DiffConfig 組み立て
        if matchers_file and os.path.isfile(matchers_file):
            config = load_config(matchers_file)
            self._log(f"マッチャー: {len(config.matchers)} 件ロード")
        else:
            config = DiffConfig()

        if include_cols:
            try:
                config.global_col_filter = parse_col_spec(include_cols)
            except Exception as e:
                self._log(f"警告: 比較列の解析エラー: {e}")

        if diff_mode == "key":
            config.key_cols = parse_col_list(key_cols_str)
            config.diff_mode = "key"
            disp = ", ".join(get_column_letter(c + 1) for c in config.key_cols)
            self._log(f"差分モード: キーJOIN  キー列: {disp}")
        else:
            config.diff_mode = "lcs"
            self._log("差分モード: LCS（行の出現順）")

        # ── 実行条件サマリ ──────────────────────────────────────────
        self._log("─" * 36)
        self._log(f"[実行条件] 旧: {old_dir}")
        self._log(f"[実行条件] 新: {new_dir}")
        if config.global_col_filter is not None:
            col_letters = sorted(get_column_letter(i + 1) for i in config.global_col_filter)
            self._log(f"[実行条件] 比較列: {', '.join(col_letters)}  (raw='{include_cols}')")
        else:
            self._log(f"[実行条件] 比較列: 全列  (raw='{include_cols}')")
        if config.diff_mode == "key":
            key_letters = ", ".join(get_column_letter(c + 1) for c in config.key_cols)
            self._log(f"[実行条件] キー列: {key_letters}")
        self._log(f"[実行条件] シート: {sheet or '全シート'}")
        self._log(f"[実行条件] ペアリング: {pairing_method}")
        self._log("─" * 36)

        # 出力フォルダ決定
        old_name = Path(old_dir).name
        new_name = Path(new_dir).name
        out_dir = output_dir or str(Path(new_dir).parent / f"{old_name}_vs_{new_name}")
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        self._log(f"出力先: {out_dir}")

        # 各ペアを比較
        results = []
        skipped = []
        for i, pair in enumerate(matched, 1):
            self._log(f"[{i}/{len(matched)}] {pair.old_name} → {pair.new_name}")
            old_path = os.path.join(old_dir, pair.old_name)
            new_path = os.path.join(new_dir, pair.new_name)

            try:
                old_sheets = read_workbook(old_path, strikethrough, sheet or None)
                new_sheets = read_workbook(new_path, strikethrough, sheet or None)

                file_diff = diff_files(
                    old_sheets, new_sheets, old_path, new_path,
                    include_strike=strikethrough, config=config,
                )

                out_name = f"{Path(pair.new_name).stem}_diff.html"
                out_path = os.path.join(out_dir, out_name)
                Path(out_path).write_text(render(file_diff), encoding="utf-8")

                results.append((pair, file_diff, out_path))

            except Exception as e:
                self._log(f"  ⚠ スキップ ({pair.old_name}): {e}")
                skipped.append(pair)

        # インデックスHTML生成
        index_path = os.path.join(out_dir, "index.html")
        Path(index_path).write_text(
            _render_index_html(results, unmatched, old_dir, new_dir),
            encoding="utf-8",
        )

        diff_count = sum(1 for _, fd, _ in results if fd.has_differences)
        nodiff_count = len(results) - diff_count
        self._log(
            f"完了: 差分あり {diff_count} 件 / 差分なし {nodiff_count} 件"
            + (f" / スキップ {len(skipped)} 件" if skipped else "")
            + (f" / 対象外 {len(unmatched)} 件" if unmatched else "")
            + f" → {index_path}"
        )

        if open_browser:
            webbrowser.open(Path(index_path).resolve().as_uri())

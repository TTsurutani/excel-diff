"""タブ③ シート分解。"""
import os
import queue
import re
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, simpledialog, ttk
from typing import Callable

from . import settings as cfg
from .widgets import FileSelectRow
from .worker import get_worker

# ファイル名として使えない文字（Windows / macOS / Linux 共通）
_INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')


def _safe_filename(name: str) -> str:
    return _INVALID_CHARS.sub("_", name)


def _apply_name_regex(sheet_name: str, pattern: re.Pattern) -> tuple[str, bool]:
    """(変換後ファイル名ベース, マッチしたか) を返す"""
    m = pattern.search(sheet_name)
    if m and m.lastindex and m.lastindex >= 1:
        return m.group(1), True
    return sheet_name, False   # フォールバック


class TabSplit(tk.Frame):

    def __init__(self, parent, log: Callable[[str], None]) -> None:
        super().__init__(parent)
        self._log = log
        self._result_q: "queue.Queue | None" = None
        self._sheet_names: list[str] = []

        self._book   = tk.StringVar(value=cfg.get("split", "book_file"))
        self._prefix = tk.StringVar(value=cfg.get("split", "prefix"))
        self._suffix = tk.StringVar(value=cfg.get("split", "suffix"))
        self._nregex = tk.StringVar(value=cfg.get("split", "name_regex"))
        self._outdir = tk.StringVar(value=cfg.get("split", "output_dir"))
        self._presets: list = []

        self._build()

        # ファイルが既に設定されていればシート名を読み込む
        if self._book.get().strip():
            self._load_sheets()

    # ================================================================== レイアウト

    def _build(self) -> None:
        pad = {"padx": 6, "pady": 3}

        # ── 入力フォーム ────────────────────────────────────────────────
        grp = tk.LabelFrame(self, text="設定")
        grp.pack(fill="x", **pad)

        FileSelectRow(
            grp, "ブックファイル", self._book,
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        ).pack(fill="x", padx=6, pady=2)
        # ファイル選択時にシート名を自動読み込み
        self._book.trace_add("write", lambda *_: self._on_book_change())

        fr_pre = tk.Frame(grp)
        fr_pre.pack(fill="x", padx=6, pady=2)
        tk.Label(fr_pre, text="前置文字列", width=14, anchor="w").pack(side="left")
        tk.Entry(fr_pre, textvariable=self._prefix, width=20).pack(side="left")

        fr_suf = tk.Frame(grp)
        fr_suf.pack(fill="x", padx=6, pady=2)
        tk.Label(fr_suf, text="後置文字列", width=14, anchor="w").pack(side="left")
        tk.Entry(fr_suf, textvariable=self._suffix, width=20).pack(side="left")

        fr_rex = tk.Frame(grp)
        fr_rex.pack(fill="x", padx=6, pady=2)
        tk.Label(fr_rex, text="名前変換正規表現", width=14, anchor="w").pack(side="left")
        tk.Entry(fr_rex, textvariable=self._nregex, width=28).pack(side="left")
        tk.Label(fr_rex, text="例: ^([^(（]+)", fg="gray").pack(side="left", padx=6)

        tk.Label(grp,
                 text="  ※ キャプチャグループ () の内容がファイル名ベースになります。空=シート名そのまま\n"
                      "  ※ ^([^(（]+) で半角・全角どちらの括弧より前も抽出できます",
                 fg="gray", font=("", 8)).pack(anchor="w", padx=6)

        fr_pst = tk.Frame(grp)
        fr_pst.pack(fill="x", padx=6, pady=2)
        tk.Label(fr_pst, text="よく使うパターン", width=14, anchor="w").pack(side="left")
        tk.Button(fr_pst, text="削除", width=5,
                  command=self._delete_preset).pack(side="right")
        tk.Button(fr_pst, text="保存", width=5,
                  command=self._save_preset).pack(side="right", padx=(2, 4))
        self._preset_combo = ttk.Combobox(fr_pst, state="readonly")
        self._preset_combo.pack(side="left", fill="x", expand=True)
        self._preset_combo.bind("<<ComboboxSelected>>", self._on_preset_selected)

        self._load_presets()

        FileSelectRow(grp, "出力先フォルダ", self._outdir, mode="dir").pack(
            fill="x", padx=6, pady=2)
        tk.Label(grp, text="  ※ 空=ブックと同じフォルダ",
                 fg="gray", font=("", 8)).pack(anchor="w", padx=6, pady=(0, 4))

        # 入力変更のたびにプレビューを更新
        for var in (self._prefix, self._suffix, self._nregex):
            var.trace_add("write", lambda *_: self._update_preview())

        # ── 正規表現エラー・一致件数表示 ───────────────────────────────
        self._lbl_regex_err = tk.Label(self, text="", fg="red", font=("", 8), anchor="w")
        self._lbl_regex_err.pack(fill="x", padx=6)
        self._lbl_match_count = tk.Label(self, text="", fg="#1a7f37", font=("", 8), anchor="w")
        self._lbl_match_count.pack(fill="x", padx=6)

        # ── プレビューテーブル ──────────────────────────────────────────
        grp_prev = tk.LabelFrame(self, text="出力プレビュー")
        grp_prev.pack(fill="both", expand=True, **pad)

        cols = ("sheet", "outfile")
        self._tree = ttk.Treeview(
            grp_prev, columns=cols, show="headings", selectmode="none",
        )
        self._tree.heading("sheet",   text="シート名")
        self._tree.heading("outfile", text="出力ファイル名")
        self._tree.column("sheet",   width=220, anchor="w")
        self._tree.column("outfile", width=320, anchor="w")
        self._tree.tag_configure("matched", foreground="#1a7f37")  # 正規表現一致・変換あり
        self._tree.tag_configure("warn",    foreground="#b58900") # 正規表現不一致（フォールバック）
        self._tree.tag_configure("error",   foreground="#cf222e") # 不正文字置換あり

        sb = ttk.Scrollbar(grp_prev, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
        sb.pack(side="left", fill="y", pady=4)

        tk.Label(grp_prev,
                 text="※ 緑=正規表現一致・変換あり\n※ 橙=正規表現不一致（シート名にフォールバック）\n※ 赤=不正文字を _ に置換",
                 fg="gray", font=("", 8), justify="left").pack(
            side="left", anchor="n", padx=6, pady=4)

        # ── 実行ボタン ──────────────────────────────────────────────────
        btn_row = tk.Frame(self)
        btn_row.pack(fill="x", padx=6, pady=(4, 6))
        self._btn_run = tk.Button(
            btn_row, text="実行", width=16,
            bg="#4a9eff", fg="white", font=("", 10, "bold"),
            command=self._run,
        )
        self._btn_run.pack(side="right")

    # ================================================================== プリセット管理

    def _load_presets(self) -> None:
        self._presets = cfg.get_split_presets()
        self._preset_combo["values"] = [p["name"] for p in self._presets]
        self._preset_combo.set("")

    def _on_preset_selected(self, _event=None) -> None:
        idx = self._preset_combo.current()
        if 0 <= idx < len(self._presets):
            self._nregex.set(self._presets[idx]["regex"])

    def _save_preset(self) -> None:
        regex = self._nregex.get().strip()
        if not regex:
            messagebox.showwarning("確認", "正規表現が入力されていません", parent=self)
            return
        name = simpledialog.askstring(
            "パターン保存", "パターン名を入力してください:", parent=self)
        if not name or not name.strip():
            return
        name = name.strip()
        for p in self._presets:
            if p["name"] == name:
                if not messagebox.askyesno(
                        "上書き確認", f"「{name}」は既に存在します。上書きしますか？", parent=self):
                    return
                p["regex"] = regex
                cfg.set_split_presets(self._presets)
                cfg.save()
                self._load_presets()
                self._preset_combo.set(name)
                return
        self._presets.append({"name": name, "regex": regex})
        cfg.set_split_presets(self._presets)
        cfg.save()
        self._load_presets()
        self._preset_combo.set(name)

    def _delete_preset(self) -> None:
        idx = self._preset_combo.current()
        if idx < 0:
            messagebox.showwarning("確認", "削除するパターンを選択してください", parent=self)
            return
        name = self._presets[idx]["name"]
        if not messagebox.askyesno("削除確認", f"「{name}」を削除しますか？", parent=self):
            return
        self._presets.pop(idx)
        cfg.set_split_presets(self._presets)
        cfg.save()
        self._load_presets()

    # ================================================================== シート名読み込み

    def _on_book_change(self) -> None:
        path = self._book.get().strip()
        if path and os.path.isfile(path) and path.lower().endswith(".xlsx"):
            self._load_sheets()
        else:
            self._sheet_names = []
            self._update_preview()

    def _load_sheets(self) -> None:
        path = self._book.get().strip()
        self._log(f"シート名読み込み中: {Path(path).name}")
        self._result_q = get_worker().submit(self._do_load_sheets, path)
        self.after(100, self._poll_load)

    def _do_load_sheets(self, path: str) -> list[str]:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    def _poll_load(self) -> None:
        if self._result_q is None:
            return
        try:
            status, val = self._result_q.get_nowait()
            if status == "err":
                self._log(f"シート読み込みエラー: {val}")
                self._sheet_names = []
            else:
                self._sheet_names = val
                self._log(f"シート数: {len(val)}")
            self._update_preview()
        except queue.Empty:
            self.after(100, self._poll_load)

    # ================================================================== プレビュー更新

    def _update_preview(self) -> None:
        if not hasattr(self, "_tree") or not hasattr(self, "_lbl_regex_err"):
            return  # まだ初期化中（安全ガード）

        for row in self._tree.get_children():
            self._tree.delete(row)
        self._lbl_regex_err.config(text="")
        self._lbl_match_count.config(text="")

        if not self._sheet_names:
            return

        prefix = self._prefix.get()
        suffix = self._suffix.get()
        nregex = self._nregex.get().strip()

        compiled: "re.Pattern | None" = None
        if nregex:
            try:
                compiled = re.compile(nregex)
                if compiled.groups < 1:
                    self._lbl_regex_err.config(
                        text="正規表現にキャプチャグループ () が必要です")
                    compiled = None
            except re.error as e:
                self._lbl_regex_err.config(text=f"正規表現エラー: {e}")
                compiled = None

        matched_count = 0
        for sheet in self._sheet_names:
            if compiled:
                base, regex_matched = _apply_name_regex(sheet, compiled)
                if regex_matched:
                    matched_count += 1
                    # 変換結果がシート名と異なる場合のみ緑で強調
                    safe_base = _safe_filename(base)
                    has_invalid = (safe_base != base)
                    out_name = f"{prefix}{safe_base}{suffix}.xlsx"
                    tag = "error" if has_invalid else "matched"
                else:
                    # 不一致：シート名のままフォールバック（橙）
                    safe_base = _safe_filename(sheet)
                    has_invalid = (safe_base != sheet)
                    out_name = f"{prefix}{safe_base}{suffix}.xlsx"
                    tag = "error" if has_invalid else "warn"
            else:
                base = sheet
                safe_base = _safe_filename(base)
                has_invalid = (safe_base != base)
                out_name = f"{prefix}{safe_base}{suffix}.xlsx"
                tag = "error" if has_invalid else ""

            self._tree.insert("", "end", values=(sheet, out_name),
                              tags=(tag,) if tag else ())

        if compiled is not None:
            total = len(self._sheet_names)
            self._lbl_match_count.config(
                text=f"正規表現が一致: {matched_count} / {total} シート"
                + ("" if matched_count > 0 else "  ← パターンがシート名と一致していません"),
            )

    # ================================================================== 状態保存

    def save_state(self) -> None:
        """現在のUI値を設定に書き戻す（ウィンドウを閉じる前に呼ばれる）。"""
        cfg.set_tab("split", {
            "book_file":  self._book.get(),
            "prefix":     self._prefix.get(),
            "suffix":     self._suffix.get(),
            "name_regex": self._nregex.get(),
            "output_dir": self._outdir.get(),
        })

    # ================================================================== 実行

    def _run(self) -> None:
        book = self._book.get().strip()
        if not book:
            messagebox.showerror("エラー", "ブックファイルを指定してください")
            return
        if not os.path.isfile(book):
            messagebox.showerror("エラー", f"ファイルが見つかりません:\n{book}")
            return
        if not book.lower().endswith(".xlsx"):
            messagebox.showerror("エラー", ".xlsx ファイルを指定してください")
            return

        nregex = self._nregex.get().strip()
        if nregex:
            try:
                pat = re.compile(nregex)
                if pat.groups < 1:
                    messagebox.showerror("エラー", "名前変換正規表現にキャプチャグループ () が必要です")
                    return
            except re.error as e:
                messagebox.showerror("エラー", f"正規表現エラー:\n{e}")
                return

        cfg.set_tab("split", {
            "book_file":  book,
            "prefix":     self._prefix.get(),
            "suffix":     self._suffix.get(),
            "name_regex": nregex,
            "output_dir": self._outdir.get(),
        })
        cfg.save()

        self._btn_run.config(state="disabled", text="実行中...")
        self._log(f"シート分解を開始: {Path(book).name}")

        self._result_q = get_worker().submit(
            self._do_split,
            book,
            self._prefix.get(),
            self._suffix.get(),
            nregex or None,
            self._outdir.get().strip() or None,
        )
        self.after(100, self._poll_run)

    def _do_split(self, book, prefix, suffix, name_regex, output_dir) -> list[str]:
        from excel_diff.splitter import split_workbook
        return split_workbook(
            book, prefix=prefix, suffix=suffix,
            name_regex=name_regex, output_dir=output_dir,
        )

    def _poll_run(self) -> None:
        if self._result_q is None:
            return
        try:
            status, val = self._result_q.get_nowait()
            self._btn_run.config(state="normal", text="実行")
            if status == "err":
                self._log(f"エラー: {val}")
            else:
                self._log(f"分解完了: {len(val)} シート")
                for p in val:
                    self._log(f"  → {p}")
        except queue.Empty:
            self.after(100, self._poll_run)

"""タブ⑤ シート比較。"""
import os
import queue
import re
import tkinter as tk
import webbrowser
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Callable

from . import settings as cfg
from .widgets import FileSelectRow
from .worker import get_worker

_EXCEL_TYPES = [("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")]


class TabSheetDiff(tk.Frame):

    def __init__(self, parent, log: Callable[[str], None]) -> None:
        super().__init__(parent)
        self._log = log
        self._result_q: "queue.Queue | None" = None
        self._old_load_q: "queue.Queue | None" = None
        self._new_load_q: "queue.Queue | None" = None

        self._old      = tk.StringVar(value=cfg.get("sheet_compare", "old_file"))
        self._new      = tk.StringVar(value=cfg.get("sheet_compare", "new_file"))
        self._out_dir  = tk.StringVar(value=cfg.get("sheet_compare", "output_dir"))
        self._cols     = tk.StringVar(value=cfg.get("sheet_compare", "include_cols"))
        self._key_cols = tk.StringVar(value=cfg.get("sheet_compare", "key_cols"))
        self._strike   = tk.BooleanVar(value=cfg.get("sheet_compare", "strikethrough"))
        self._open_br  = tk.BooleanVar(value=cfg.get("sheet_compare", "open_browser"))
        self._mode     = tk.StringVar(value=cfg.get("sheet_compare", "diff_mode", "lcs"))

        self._build()

        self._old.trace_add("write", lambda *_: self._on_file_change("old"))
        self._new.trace_add("write", lambda *_: self._on_file_change("new"))

        # 起動時に既存パスがあればシートを読み込む
        for side in ("old", "new"):
            self._on_file_change(side)

    # ------------------------------------------------------------------ レイアウト

    def _build(self) -> None:
        pad = {"padx": 6, "pady": 3}

        # 差分モード
        grp_mode = tk.LabelFrame(self, text="差分モード")
        grp_mode.pack(fill="x", **pad)

        tk.Radiobutton(
            grp_mode, text="LCS（行の出現順）",
            variable=self._mode, value="lcs", command=self._on_mode,
        ).grid(row=0, column=0, sticky="w", padx=6, pady=2)

        fr_key = tk.Frame(grp_mode)
        fr_key.grid(row=1, column=0, sticky="w", padx=6, pady=2)
        tk.Radiobutton(
            fr_key, text="キーJOIN",
            variable=self._mode, value="key", command=self._on_mode,
        ).pack(side="left")
        tk.Label(fr_key, text="キー列").pack(side="left", padx=(8, 2))
        self._entry_key = tk.Entry(fr_key, textvariable=self._key_cols, width=14)
        self._entry_key.pack(side="left")
        tk.Label(fr_key, text="（例: B  または  B,C）", foreground="gray").pack(side="left", padx=4)

        # 比較オプション（旧/新ファイルとシート選択）
        grp_files = tk.LabelFrame(self, text="比較オプション")
        grp_files.pack(fill="both", expand=True, **pad)

        FileSelectRow(
            grp_files, "出力フォルダ", self._out_dir, mode="dir",
        ).pack(fill="x", padx=6, pady=(6, 2))

        self._lb_old = self._make_file_block(grp_files, "old", "旧ファイル")
        self._lb_new = self._make_file_block(grp_files, "new", "新ファイル")

        # オプション
        grp_opt = tk.LabelFrame(self, text="オプション")
        grp_opt.pack(fill="x", **pad)

        fr_cols = tk.Frame(grp_opt)
        fr_cols.pack(fill="x", padx=6, pady=2)
        tk.Label(fr_cols, text="比較列", width=14, anchor="w").pack(side="left")
        tk.Entry(fr_cols, textvariable=self._cols).pack(side="left", fill="x", expand=True)
        tk.Label(fr_cols, text="（例: A:C,E）", foreground="gray").pack(side="left", padx=4)

        tk.Checkbutton(
            grp_opt, text="取り消し線も差分として扱う", variable=self._strike,
        ).pack(anchor="w", padx=6, pady=1)
        tk.Checkbutton(
            grp_opt, text="完了後ブラウザで開く", variable=self._open_br,
        ).pack(anchor="w", padx=6, pady=(1, 4))

        # 実行ボタン
        fr_btn = tk.Frame(self)
        fr_btn.pack(fill="x", padx=6, pady=4)
        self._btn_run = tk.Button(fr_btn, text="実行", width=12, command=self._run)
        self._btn_run.pack(side="right")

        self._on_mode()

    def _make_file_block(self, parent: tk.Widget, side: str, label: str) -> tk.Listbox:
        """FileSelectRow + シート選択Listbox を parent に pack し Listbox を返す。"""
        var = self._old if side == "old" else self._new

        FileSelectRow(parent, label, var, filetypes=_EXCEL_TYPES).pack(
            fill="x", padx=6, pady=(6, 2)
        )

        fr_lb = tk.Frame(parent)
        fr_lb.pack(fill="x", padx=6, pady=(0, 6))
        tk.Label(fr_lb, text="シート選択", width=14, anchor="w").pack(side="left", anchor="n")

        fr_inner = tk.Frame(fr_lb)
        fr_inner.pack(side="left", fill="x", expand=True)
        lb = tk.Listbox(fr_inner, height=5, selectmode="single", exportselection=False)
        sb = ttk.Scrollbar(fr_inner, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        lb.pack(side="left", fill="x", expand=True)
        sb.pack(side="left", fill="y")

        return lb

    # ------------------------------------------------------------------ モード切替

    def _on_mode(self) -> None:
        state = "normal" if self._mode.get() == "key" else "disabled"
        self._entry_key.config(state=state)

    # ------------------------------------------------------------------ シート読み込み

    def _on_file_change(self, side: str) -> None:
        lb = self._lb_old if side == "old" else self._lb_new
        lb.delete(0, "end")
        path = (self._old if side == "old" else self._new).get().strip()
        if path and os.path.isfile(path) and path.lower().endswith((".xlsx", ".xlsm")):
            q = get_worker().submit(self._do_load_sheets, path)
            if side == "old":
                self._old_load_q = q
            else:
                self._new_load_q = q
            self.after(100, lambda: self._poll_load(side, q))

    def _do_load_sheets(self, path: str) -> list:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names

    def _poll_load(self, side: str, q: queue.Queue) -> None:
        try:
            status, val = q.get_nowait()
            lb = self._lb_old if side == "old" else self._lb_new
            lb.delete(0, "end")
            if status == "err":
                self._log(f"シート読み込みエラー ({side}): {val}")
            else:
                for name in val:
                    lb.insert("end", name)
                saved = cfg.get("sheet_compare", f"{side}_sheet", "")
                if saved:
                    items = lb.get(0, "end")
                    if saved in items:
                        idx = list(items).index(saved)
                        lb.selection_set(idx)
                        lb.see(idx)
        except queue.Empty:
            self.after(100, lambda: self._poll_load(side, q))

    # ------------------------------------------------------------------ 実行

    def _run(self) -> None:
        old_file = self._old.get().strip()
        new_file = self._new.get().strip()

        if not old_file or not new_file:
            messagebox.showerror("エラー", "旧ファイルと新ファイルを指定してください")
            return
        for path, label in ((old_file, "旧"), (new_file, "新")):
            if not os.path.isfile(path):
                messagebox.showerror("エラー", f"{label}ファイルが見つかりません:\n{path}")
                return

        sel_old = self._lb_old.curselection()
        sel_new = self._lb_new.curselection()
        if not sel_old:
            messagebox.showerror("エラー", "旧ファイルのシートを選択してください")
            return
        if not sel_new:
            messagebox.showerror("エラー", "新ファイルのシートを選択してください")
            return
        old_sheet = self._lb_old.get(sel_old[0])
        new_sheet = self._lb_new.get(sel_new[0])

        if self._mode.get() == "key" and not self._key_cols.get().strip():
            messagebox.showerror("エラー", "キーJOINモード: キー列を指定してください")
            return

        self._btn_run.config(state="disabled", text="実行中...")
        self._result_q = get_worker().submit(
            self._do_diff,
            old_file, new_file, old_sheet, new_sheet,
            self._out_dir.get().strip(),
            self._cols.get().strip(),
            self._strike.get(),
            self._mode.get(),
            self._key_cols.get().strip(),
            self._open_br.get(),
        )
        self.after(100, self._poll)

    def _poll(self) -> None:
        if self._result_q is None:
            return
        try:
            status, val = self._result_q.get_nowait()
            self._btn_run.config(state="normal", text="実行")
            if status == "err":
                self._log(f"エラー: {val}")
        except queue.Empty:
            self.after(100, self._poll)

    def _do_diff(
        self, old_file, new_file, old_sheet, new_sheet,
        output_dir, include_cols, strikethrough, diff_mode, key_cols_str, open_browser,
    ) -> None:
        from excel_diff.reader import read_workbook
        from excel_diff.diff_engine import diff_files, RowTag
        from excel_diff.html_renderer import render
        from excel_diff.matcher import DiffConfig, parse_col_spec, parse_col_list
        from openpyxl.utils import get_column_letter

        self._log(f"読み込み中: {Path(old_file).name}  [{old_sheet}]")
        old_sheets = read_workbook(old_file, strikethrough, sheet_filter=old_sheet)
        if old_sheet not in old_sheets:
            self._log(f"エラー: シート '{old_sheet}' が見つかりません（旧ファイル）")
            return

        self._log(f"読み込み中: {Path(new_file).name}  [{new_sheet}]")
        new_sheets = read_workbook(new_file, strikethrough, sheet_filter=new_sheet)
        if new_sheet not in new_sheets:
            self._log(f"エラー: シート '{new_sheet}' が見つかりません（新ファイル）")
            return

        # シート名が異なる場合でも diff_files が同一シートとして扱えるよう new 側のキーを統一
        old_data = {old_sheet: old_sheets[old_sheet]}
        new_data = {old_sheet: new_sheets[new_sheet]}

        config = DiffConfig()
        if include_cols:
            try:
                config.global_col_filter = parse_col_spec(include_cols)
            except Exception as e:
                self._log(f"警告: 比較列の解析エラー: {e}")

        if diff_mode == "key":
            config.key_cols = parse_col_list(key_cols_str)
            config.diff_mode = "key"
            disp = ", ".join(get_column_letter(c + 1) for c in config.key_cols)
            self._log(f"差分モード: キーJOIN  キー列: {disp}")
        else:
            config.diff_mode = "lcs"
            self._log("差分モード: LCS（行の出現順）")

        self._log("差分計算中...")
        old_label = f"{Path(old_file).name} [{old_sheet}]"
        new_label = f"{Path(new_file).name} [{new_sheet}]"
        file_diff = diff_files(
            old_data, new_data, old_label, new_label,
            include_strike=strikethrough, config=config,
        )

        def _safe(name: str) -> str:
            return re.sub(r'[\\/:*?"<>|]', "_", name)

        base_dir = Path(output_dir) if output_dir and os.path.isdir(output_dir) else Path(new_file).parent
        out_path = str(base_dir / f"{Path(new_file).stem}_{_safe(old_sheet)}_vs_{_safe(new_sheet)}.html")
        Path(out_path).write_text(render(file_diff), encoding="utf-8")

        if file_diff.has_differences:
            delete = sum(1 for sd in file_diff.sheet_diffs
                         for rd in sd.row_diffs if rd.tag == RowTag.DELETE)
            insert = sum(1 for sd in file_diff.sheet_diffs
                         for rd in sd.row_diffs if rd.tag == RowTag.INSERT)
            modify = sum(1 for sd in file_diff.sheet_diffs
                         for rd in sd.row_diffs if rd.tag == RowTag.MODIFY)
            self._log(
                f"差分あり (削除 {delete} 追加 {insert} 変更 {modify})"
                f"  [{old_sheet}] ↔ [{new_sheet}]"
                f"  → {out_path}"
            )
        else:
            self._log(f"差分なし  [{old_sheet}] ↔ [{new_sheet}]")

        if open_browser:
            webbrowser.open(Path(out_path).resolve().as_uri())

    # ------------------------------------------------------------------ 設定保存

    def save_state(self) -> None:
        sel_old = self._lb_old.curselection()
        sel_new = self._lb_new.curselection()
        cfg.set_tab("sheet_compare", {
            "old_file":      self._old.get(),
            "new_file":      self._new.get(),
            "old_sheet":     self._lb_old.get(sel_old[0]) if sel_old else "",
            "new_sheet":     self._lb_new.get(sel_new[0]) if sel_new else "",
            "output_dir":    self._out_dir.get(),
            "include_cols":  self._cols.get(),
            "strikethrough": self._strike.get(),
            "open_browser":  self._open_br.get(),
            "diff_mode":     self._mode.get(),
            "key_cols":      self._key_cols.get(),
        })
